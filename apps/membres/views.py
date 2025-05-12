# apps/membres/views.py
import csv
import io
import json
import logging
from datetime import datetime
from django.conf import settings
from django.db.models.functions import ExtractMonth
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q, F, IntegerField, Case, When
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import (
    CreateView, DeleteView, DetailView, FormView, ListView, TemplateView, UpdateView, View
)

from openpyxl import Workbook
from openpyxl import load_workbook

from apps.core.mixins import StaffRequiredMixin, TrashViewMixin, RestoreViewMixin
from apps.core.models import Statut
from apps.membres.forms import (
    MembreForm, TypeMembreForm, MembreTypeMembreForm, 
    MembreImportForm, MembreSearchForm
)
from apps.membres.models import Membre, TypeMembre, MembreTypeMembre, HistoriqueMembre
from django.db.models import F, IntegerField
from django.utils.crypto import get_random_string
from apps.accounts.models import CustomUser
from django.http import Http404
import types
import openpyxl.styles

logger = logging.getLogger(__name__)


class DashboardView(TemplateView):
    """
    Vue du tableau de bord pour les statistiques sur les membres
    """
    template_name = 'membres/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Récupérer la date actuelle et les statistiques de base
        today = timezone.now().date()
        total_membres = Membre.objects.count()
        membres_actifs = Membre.objects.actifs().count()
        
        # Types de membres avec leur nombre
        types_membres = TypeMembre.objects.annotate(
            count=Count('membres_historique__membre', 
                          filter=Q(membres_historique__date_debut__lte=today, 
                                 membres_historique__date_fin__isnull=True),
                          distinct=True)
        ).order_by('-count')
        
        # Adhésions récentes
        adhesions_recentes = Membre.objects.order_by('-date_adhesion')[:10]
        
        # Statistiques sur les membres
        membres_par_mois = Membre.objects.annotate(
        month=ExtractMonth('date_adhesion', output_field=IntegerField())
        ).values('month').annotate(count=Count('id')).order_by('month')
        
        # Membres par statut - Inclure l'ID du statut dans les valeurs retournées
        membres_par_statut = Membre.objects.filter(
            statut__isnull=False
        ).values('statut__nom', 'statut_id').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Membres avec/sans compte utilisateur
        avec_compte = Membre.objects.filter(utilisateur__isnull=False).count()
        sans_compte = total_membres - avec_compte
        
        # Construire les données pour les graphiques
        chart_types = [
            {'id': t.id, 'name': t.libelle, 'value': t.count} for t in types_membres
        ]
        
        chart_monthly = []
        for i in range(1, 13):
            month_name = datetime(2000, i, 1).strftime('%B')
            count = next((item['count'] for item in membres_par_mois if item['month'] == i), 0)
            chart_monthly.append({'month': month_name, 'count': count})
        
        # Corriger l'accès aux données du statut (utiliser la syntaxe dictionnaire)
        chart_statuts = [
            {'id': s['statut_id'], 'name': s['statut__nom'] or 'Sans statut', 'value': s['count']} 
            for s in membres_par_statut
        ]
        
        chart_comptes = [
            {'name': 'Avec compte', 'value': avec_compte},
            {'name': 'Sans compte', 'value': sans_compte}
        ]
        
        context.update({
            'total_membres': total_membres,
            'membres_actifs': membres_actifs,
            'types_membres': types_membres,
            'adhesions_recentes': adhesions_recentes,
            'chart_types': json.dumps(chart_types),
            'chart_monthly': json.dumps(chart_monthly),
            'chart_statuts': json.dumps(chart_statuts),
            'chart_comptes': json.dumps(chart_comptes),
        })
        
        return context

class MembreListView(ListView):
    """
    Vue pour afficher la liste des membres avec recherche et filtrage
    """
    model = Membre
    template_name = 'membres/liste.html'
    context_object_name = 'membres'
    paginate_by = 20
    
    def get_queryset(self):
        """
        Méthode pour filtrer les membres selon les critères de recherche.
        """
        # Commencer avec tous les membres (non supprimés)
        queryset = Membre.objects.all()
        
        # Récupérer et valider le formulaire
        form = MembreSearchForm(self.request.GET)
        
        if not form.is_valid():
            return queryset
            
        # Récupérer les paramètres de tri
        sort_by = self.request.GET.get('sort', 'nom')
        sort_dir = self.request.GET.get('dir', 'asc')
        
        # Variables pour stocker les différents critères de filtrage
        term = form.cleaned_data.get('terme')
        type_membre = form.cleaned_data.get('type_membre')
        type_membre_id = type_membre.id if type_membre else None
        statut = form.cleaned_data.get('statut')
        statut_id = statut.id if statut else None
        date_adhesion_min = form.cleaned_data.get('date_adhesion_min')
        date_adhesion_max = form.cleaned_data.get('date_adhesion_max')
        age_min = form.cleaned_data.get('age_min')
        age_max = form.cleaned_data.get('age_max')
        cotisations_impayees = form.cleaned_data.get('cotisations_impayees')
        avec_compte = form.cleaned_data.get('avec_compte')
        actif = form.cleaned_data.get('actif')
        
        # Filtres de date d'adhésion
        if date_adhesion_min:
            queryset = queryset.filter(date_adhesion__gte=date_adhesion_min)
            
        if date_adhesion_max:
            queryset = queryset.filter(date_adhesion__lte=date_adhesion_max)
        
        # Filtre par terme de recherche
        if term:
            q_objects = (
                Q(nom__icontains=term) | 
                Q(prenom__icontains=term) | 
                Q(email__icontains=term) | 
                Q(telephone__icontains=term) |
                Q(code_postal__icontains=term) |
                Q(ville__icontains=term)
            )
            queryset = queryset.filter(q_objects)
        
        # Filtre par type de membre
        if type_membre_id:
            queryset = queryset.filter(
                types_historique__type_membre_id=type_membre_id,
                types_historique__date_debut__lte=timezone.now().date(),
                types_historique__date_fin__isnull=True
            ).distinct()
        
        # Filtre par statut
        if statut_id:
            queryset = queryset.filter(statut_id=statut_id)
        
        # Filtres par âge
        if age_min is not None or age_max is not None:
            today = timezone.now().date()
            
            if age_min is not None:
                date_naissance_max = today.replace(year=today.year - age_min)
                queryset = queryset.filter(date_naissance__lte=date_naissance_max)
                
            if age_max is not None:
                date_naissance_min = today.replace(year=today.year - age_max - 1)
                date_naissance_min = date_naissance_min.replace(day=date_naissance_min.day + 1)
                queryset = queryset.filter(date_naissance__gt=date_naissance_min)

        # Filtre par cotisations impayées
        if cotisations_impayees:
            try:
                from apps.cotisations.models import Cotisation
                
                # Récupérer les valeurs uniques de statut_paiement
                statuts_uniques = list(Cotisation.objects.values_list('statut_paiement', flat=True).distinct())
                
                # Approche simple mais efficace :
                # 1. Trouver les IDs des membres qui ont au moins une cotisation impayée
                from django.db.models import Count, Case, When, IntegerField
                
                # Cette requête annotée compte les cotisations impayées pour chaque membre
                membres_avec_comptage = Cotisation.objects.values('membre_id').annotate(
                    nb_impayees=Count(
                        Case(
                            When(~Q(statut_paiement='payée'), then=1),
                            output_field=IntegerField()
                        )
                    )
                ).filter(nb_impayees__gt=0)
                
                # Extraire uniquement les IDs des membres
                membres_avec_impayees = [item['membre_id'] for item in membres_avec_comptage]
                
                if membres_avec_impayees:
                    # Filtrer le queryset pour ne garder que les membres avec cotisations impayées
                    queryset = queryset.filter(id__in=membres_avec_impayees)
                else:
                    # Aucune cotisation impayée trouvée
                    queryset = queryset.none()  # Retourne un queryset vide
                    
            except ImportError as e:
                # Le module cotisations n'est pas disponible
                pass
            except Exception as e:
                # Capturer toute autre exception pour éviter l'échec silencieux
                pass
        
        # Filtre par compte utilisateur
        if avec_compte == 'avec':
            queryset = queryset.filter(utilisateur__isnull=False)
        elif avec_compte == 'sans':
            queryset = queryset.filter(utilisateur__isnull=True)
        
        # Filtre par statut d'activité
        if actif == 'actif':
            # Obtenir les IDs des membres avec au moins un type de membre actif
            membres_ids = MembreTypeMembre.objects.filter(
                date_debut__lte=timezone.now().date(),
                date_fin__isnull=True
            ).values_list('membre_id', flat=True).distinct()
            
            queryset = queryset.filter(id__in=membres_ids)
        elif actif == 'inactif':
            # Obtenir les IDs des membres avec au moins un type de membre actif
            membres_ids = MembreTypeMembre.objects.filter(
                date_debut__lte=timezone.now().date(),
                date_fin__isnull=True
            ).values_list('membre_id', flat=True).distinct()
            
            queryset = queryset.exclude(id__in=membres_ids)
        
        # Appliquer le tri
        if sort_by:
            direction = '' if sort_dir == 'asc' else '-'
            
            if sort_by == 'nom':
                order_fields = [f'{direction}nom', f'{direction}prenom']
            elif sort_by == 'email':
                order_fields = [f'{direction}email']
            elif sort_by == 'telephone':
                order_fields = [f'{direction}telephone']
            elif sort_by == 'date_adhesion':
                order_fields = [f'{direction}date_adhesion']
            elif sort_by == 'statut':
                queryset = queryset.select_related('statut')
                order_fields = [f'{direction}statut__nom', f'{direction}nom']
            elif sort_by == 'types':
                queryset = queryset.annotate(
                    nb_types=Count(
                        'types_historique',
                        filter=Q(
                            types_historique__date_debut__lte=timezone.now().date(),
                            types_historique__date_fin__isnull=True
                        ),
                        distinct=True
                    )
                )
                order_fields = [f'{direction}nb_types', f'{direction}nom']
            else:
                order_fields = ['nom', 'prenom']
            
            queryset = queryset.order_by(*order_fields)
        
        # Précharger les relations pour optimiser les performances
        result = queryset.select_related('statut').prefetch_related('types')
        
        return result
    

    # Ajouter à la méthode get_context_data pour passer les paramètres de tri au template
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Ajouter le formulaire de recherche
        context['search_form'] = MembreSearchForm(self.request.GET)
        
        # Statistiques rapides
        context['total_membres'] = Membre.objects.count()
        context['membres_actifs'] = Membre.objects.actifs().count()
        
        # Paramètres de tri pour le template
        context['sort'] = self.request.GET.get('sort', 'nom')
        context['sort_dir'] = self.request.GET.get('dir', 'asc')
        
        # Créer les paramètres de requête pour les liens de tri
        # en excluant les paramètres de tri existants
        query_params = self.request.GET.copy()
        if 'sort' in query_params:
            query_params.pop('sort')
        if 'dir' in query_params:
            query_params.pop('dir')
        context['query_params'] = query_params.urlencode()
        
        return context

class MembreDetailView(DetailView):
    """
    Vue pour afficher les détails d'un membre
    """
    model = Membre
    template_name = 'membres/detail.html'
    context_object_name = 'membre'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        membre = self.object
        
        # Types de membre actuels et historiques
        context['types_actifs'] = membre.get_types_actifs()
        context['historique_types'] = MembreTypeMembre.objects.filter(
            membre=membre
        ).select_related('type_membre', 'modifie_par').order_by('-date_debut')
        
        # Formulaire pour ajouter un nouveau type
        context['type_form'] = MembreTypeMembreForm(membre=membre, user=self.request.user)
        
        # Historique des modifications
        context['historique'] = HistoriqueMembre.objects.filter(
            membre=membre
        ).select_related('utilisateur').order_by('-created_at')[:10]
        
        # Informations supplémentaires calculées
        context['age'] = membre.age()
        context['anciennete'] = (timezone.now().date() - membre.date_adhesion).days // 365
        
        # Cotisations (si l'application est disponible)
        try:
            from apps.cotisations.models import Cotisation
            context['cotisations'] = Cotisation.objects.filter(
                membre=membre
            ).order_by('-annee', '-mois')[:5]
            context['nb_cotisations_impayees'] = Cotisation.objects.filter(
                membre=membre, 
                statut_paiement__in=['non_payée', 'partiellement_payée']
            ).count()
        except ImportError:
            context['cotisations'] = None
        
        # Événements (si l'application est disponible)
        try:
            from apps.evenements.models import Inscription
            context['inscriptions'] = Inscription.objects.filter(
                membre=membre
            ).select_related('evenement').order_by('-date_inscription')[:5]
        except ImportError:
            context['inscriptions'] = None
        
        # Documents (si l'application est disponible)
        try:
            from apps.documents.models import Document
            context['documents'] = Document.objects.filter(
                reference_id=membre.id,
                type_document__nom='membre'
            ).order_by('-date_upload')
        except ImportError:
            context['documents'] = None
        
        return context

class MembreCreateView(StaffRequiredMixin, CreateView):
    """
    Vue pour créer un nouveau membre
    """
    model = Membre
    form_class = MembreForm
    template_name = 'membres/form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_invalid(self, form):
        """Affichage détaillé des erreurs de validation pour l'utilisateur"""
        for field, errors in form.errors.items():
            for error in errors:
                field_name = field if field == '__all__' else form[field].label or field
                messages.error(self.request, f"Erreur dans {field_name}: {error}")
        
        return super().form_invalid(form)
    
    def form_valid(self, form):
        try:
            # Enregistrer le membre
            membre = form.save()
            
            # Créer un compte utilisateur si demandé
            if form.cleaned_data.get('creer_compte'):
                username = f"{membre.prenom.lower()}.{membre.nom.lower()}".replace(' ', '_')
                base_username = username
                counter = 1
                
                # Éviter les doublons
                while CustomUser.objects.filter(username=username).exists():
                    username = f"{base_username}{counter}"
                    counter += 1
                
                # Utiliser le mot de passe fourni ou en générer un
                password = form.cleaned_data.get('password')
                if not password:
                    password = get_random_string(length=12)
                
                # Créer l'utilisateur
                user = CustomUser.objects.create_user(
                    username=username,
                    email=membre.email,
                    password=password,
                    first_name=membre.prenom,
                    last_name=membre.nom,
                    password_temporary=True
                )
                
                # Lier à ce membre
                membre.utilisateur = user
                membre.save(update_fields=['utilisateur'])
                
                # Vérification d'email si nécessaire
                if hasattr(settings, 'ACCOUNT_EMAIL_VERIFICATION_REQUIRED') and settings.ACCOUNT_EMAIL_VERIFICATION_REQUIRED:
                    try:
                        # Générer et envoyer le token de vérification
                        token = user.generate_email_verification_token()
                        verification_url = self.request.build_absolute_uri(
                            reverse('email_verify', kwargs={'token': token})
                        )
                        
                        from apps.core.services import EmailService
                        EmailService.send_template_email(
                            'emails/email_verification',
                            {'user': user, 'verification_url': verification_url},
                            _("Vérification de votre adresse email"),
                            user.email
                        )
                        
                        messages.info(
                            self.request,
                            _("Un email de vérification a été envoyé à %(email)s.") % {'email': user.email}
                        )
                    except Exception as email_error:
                        logger.error(f"Erreur lors de l'envoi de l'email de vérification: {str(email_error)}")
                        messages.warning(
                            self.request,
                            _("L'email de vérification n'a pas pu être envoyé.")
                        )
                else:
                    # Envoyer un email de bienvenue avec les informations de connexion
                    try:
                        from django.core.mail import EmailMultiAlternatives
                        from django.template.loader import render_to_string
                        
                        # Contexte pour le template d'email
                        context = {
                            'membre': membre,
                            'username': username,
                            'password': password,
                            'login_url': self.request.build_absolute_uri(reverse('accounts:login')),
                        }
                        
                        # Rendre les templates HTML et texte
                        html_message = render_to_string('emails/nouveau_compte.html', context)
                        text_message = render_to_string('emails/nouveau_compte.txt', context)
                        
                        # Envoyer l'email
                        email = EmailMultiAlternatives(
                            _("Bienvenue à l'association - Vos identifiants de connexion"),
                            text_message,
                            settings.DEFAULT_FROM_EMAIL,
                            [membre.email]
                        )
                        email.attach_alternative(html_message, "text/html")
                        email.send()
                        
                        logger.info(f"Email d'identifiants envoyé à {membre.email}")
                    except Exception as e:
                        logger.error(f"Erreur lors de l'envoi de l'email de bienvenue: {str(e)}")
                        messages.warning(
                            self.request,
                            _("Le membre a été créé, mais l'envoi de l'email avec les identifiants a échoué.")
                        )
                
                messages.success(
                    self.request,
                    _("Le membre %(nom)s a été créé avec succès avec un compte utilisateur (%(username)s).") % 
                    {'nom': membre.nom_complet, 'username': username}
                )
            else:
                messages.success(
                    self.request,
                    _("Le membre %(nom)s a été créé avec succès.") % {'nom': membre.nom_complet}
                )
            
            # Ajouter un enregistrement dans l'historique
            HistoriqueMembre.objects.create(
                membre=membre,
                utilisateur=self.request.user,
                action='creation',
                description=_("Création du membre"),
                donnees_apres={
                    field: str(value) for field, value in form.cleaned_data.items() 
                    if field not in ['types_membre', 'photo', 'creer_compte', 'password', 'password_confirm']
                }
            )
            
            return redirect(membre.get_absolute_url())
        except Exception as e:
            logger.error(f"Erreur lors de la création d'un membre: {str(e)}", exc_info=True)
            messages.error(self.request, _("Erreur lors de la création du membre: %(error)s") % {'error': str(e)})
            return self.form_invalid(form)

class MembreDeleteView(StaffRequiredMixin, DeleteView):
    """
    Vue pour supprimer un membre
    """
    model = Membre
    template_name = 'membres/confirmer_suppression.html'
    success_url = reverse_lazy('membres:membre_liste')
    
    def delete(self, request, *args, **kwargs):
        membre = self.get_object()
        
        # Enregistrer l'action dans l'historique
        HistoriqueMembre.objects.create(
            membre=membre,
            utilisateur=request.user,
            action='suppression',
            description=_("Suppression du membre"),
            donnees_avant={
                'nom': membre.nom,
                'prenom': membre.prenom,
                'email': membre.email,
                'date_adhesion': str(membre.date_adhesion)
            }
        )
        
        # Suppression logique explicite
        membre.deleted_at = timezone.now()
        membre.save(update_fields=['deleted_at'])
        
        messages.success(
            request, 
            _("Le membre %(name)s a été supprimé avec succès.") % {'name': membre.nom_complet}
        )
        return redirect(self.success_url)


class TypeMembreListView(ListView):
    """
    Vue pour afficher la liste des types de membres
    """
    model = TypeMembre
    template_name = 'membres/type_membre_liste.html'
    context_object_name = 'types_membres'
    
    def get_queryset(self):
        # Utiliser une annotation pour ajouter members_count à chaque type
        today = timezone.now().date()
        return TypeMembre.objects.annotate(
            members_count=Count(
                'membres_historique__membre',
                filter=Q(
                    membres_historique__date_debut__lte=today,
                    membres_historique__date_fin__isnull=True
                ),
                distinct=True
            )
        ).order_by('ordre_affichage', 'libelle')


class TypeMembreCreateView(StaffRequiredMixin, CreateView):
    """
    Vue pour créer un nouveau type de membre
    """
    model = TypeMembre
    form_class = TypeMembreForm
    template_name = 'membres/type_membre_form.html'
    success_url = reverse_lazy('membres:type_membre_liste')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            _("Le type de membre %(name)s a été créé avec succès.") % {'name': form.instance.libelle}
        )
        return response


class TypeMembreUpdateView(StaffRequiredMixin, UpdateView):
    """
    Vue pour modifier un type de membre existant
    """
    model = TypeMembre
    form_class = TypeMembreForm
    template_name = 'membres/type_membre_form.html'
    success_url = reverse_lazy('membres:type_membre_liste')
    
    def get_object(self, queryset=None):
        """
        Récupérer l'objet sans tenter d'annoter ou de compter les membres,
        ce qui peut causer des erreurs avec les relations.
        """
        if queryset is None:
            queryset = self.get_queryset()
        
        # Utiliser la méthode standard pour récupérer l'objet par pk
        pk = self.kwargs.get(self.pk_url_kwarg)
        queryset = queryset.filter(pk=pk)
        
        try:
            # Get the single item from the filtered queryset
            obj = queryset.get()
        except queryset.model.DoesNotExist:
            raise Http404(_("Aucun type de membre trouvé avec cet identifiant"))
            
        # Ajouter dynamiquement une méthode nb_membres_actifs à l'objet
        # pour être compatible avec le template existant
        type_membre = obj
        today = timezone.now().date()
        
        # Calculer le nombre de membres actifs
        membres_count = type_membre.membres_historique.filter(
            date_debut__lte=today,
            date_fin__isnull=True
        ).values('membre').distinct().count()
        
        # Ajouter une méthode dynamique à l'objet qui renvoie ce nombre
        # Cette approche évite de modifier le template
        def nb_membres_actifs(self):
            return membres_count
            
        # Lier la méthode à l'instance
        import types
        type_membre.nb_membres_actifs = types.MethodType(nb_membres_actifs, type_membre)
        
        return obj
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            _("Le type de membre %(name)s a été modifié avec succès.") % {'name': form.instance.libelle}
        )
        return response


class TypeMembreDeleteView(StaffRequiredMixin, DeleteView):
    """
    Vue pour supprimer un type de membre
    """
    model = TypeMembre
    template_name = 'membres/type_membre_confirmer_suppression.html'
    success_url = reverse_lazy('membres:type_membre_liste')
    context_object_name = 'type_membre'  # Ajouter cette ligne pour s'assurer que l'objet est dans le contexte
    
    def get_object(self, queryset=None):
        """
        Récupérer l'objet sans tenter d'annoter ou de compter les membres,
        ce qui peut causer des erreurs avec les relations.
        """
        if queryset is None:
            queryset = self.get_queryset()
        
        # Utilisez self.pk_url_kwarg pour la cohérence et la flexibilité
        pk = self.kwargs.get(self.pk_url_kwarg)
        queryset = queryset.filter(pk=pk)
        
        try:
            obj = queryset.get()
        except queryset.model.DoesNotExist:
            raise Http404(_("Aucun type de membre trouvé avec cet identifiant"))
            
        # Ajouter dynamiquement une méthode nb_membres_actifs à l'objet
        # pour être compatible avec le template existant
        type_membre = obj
        today = timezone.now().date()
        
        # Calculer le nombre de membres actifs
        membres_count = type_membre.membres_historique.filter(
            date_debut__lte=today,
            date_fin__isnull=True
        ).values('membre').distinct().count()
        
        # Ajouter une méthode dynamique à l'objet qui renvoie ce nombre
        def nb_membres_actifs(self):
            return membres_count
            
        # Lier la méthode à l'instance
        type_membre.nb_membres_actifs = types.MethodType(nb_membres_actifs, type_membre)
        
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # S'assurer que nous avons une valeur pour nb_membres_actifs dans le contexte
        context['nb_membres_actifs'] = self.object.nb_membres_actifs()
        
        # Ajouter explicitement la clé primaire au contexte
        context['object_id'] = self.object.pk
        
        return context
    
    def delete(self, request, *args, **kwargs):
        type_membre = self.get_object()
        
        # Suppression logique (soft delete)
        type_membre.delete()
        
        messages.success(
            request, 
            _("Le type de membre %(name)s a été supprimé avec succès.") % {'name': type_membre.libelle}
        )
        return redirect(self.success_url)


class MembreTypeMembreCreateView(StaffRequiredMixin, CreateView):
    """
    Vue pour ajouter un type de membre à un membre
    """
    model = MembreTypeMembre
    form_class = MembreTypeMembreForm
    template_name = 'membres/membre_type_membre_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        self.membre = get_object_or_404(Membre, pk=self.kwargs['membre_id'])
        kwargs['membre'] = self.membre
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['membre'] = self.membre
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            _("Le type de membre %(type)s a été ajouté avec succès à %(name)s.") % {
                'type': form.instance.type_membre.libelle,
                'name': form.instance.membre.nom_complet
            }
        )
        return response
    
    def get_success_url(self):
        return reverse('membres:membre_detail', kwargs={'pk': self.membre.pk})


class MembreTypeMembreUpdateView(StaffRequiredMixin, UpdateView):
    """
    Vue pour modifier l'association entre un membre et un type de membre
    """
    model = MembreTypeMembre
    form_class = MembreTypeMembreForm
    template_name = 'membres/membre_type_membre_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['membre'] = self.object.membre
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['membre'] = self.object.membre
        context['is_update'] = True
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            _("L'association de type de membre a été modifiée avec succès.")
        )
        return response
    
    def get_success_url(self):
        return reverse('membres:membre_detail', kwargs={'pk': self.object.membre.pk})


class MembreTypeMembreTerminerView(StaffRequiredMixin, View):
    """
    Vue pour terminer l'association entre un membre et un type de membre
    """
    def post(self, request, pk):
        association = get_object_or_404(MembreTypeMembre, pk=pk)
        membre = association.membre
        
        # Date de fin par défaut aujourd'hui
        date_fin = timezone.now().date()
        commentaire = request.POST.get('commentaire', '')
        
        # Mettre à jour l'association
        association.date_fin = date_fin
        association.commentaire = commentaire
        association.modifie_par = request.user
        association.save()
        
        messages.success(
            request, 
            _("L'association avec le type %(type)s a été terminée.") % {
                'type': association.type_membre.libelle
            }
        )
        
        return redirect('membres:membre_detail', pk=membre.pk)


class MembreImportView(StaffRequiredMixin, FormView):
    """
    Vue pour importer des membres depuis un fichier CSV ou Excel
    """
    form_class = MembreImportForm
    template_name = 'membres/import.html'
    success_url = reverse_lazy('membres:membre_liste')
    
    def form_valid(self, form):
        fichier = form.cleaned_data['fichier']
        extension = fichier.name.split('.')[-1].lower()
        
        type_membre = form.cleaned_data['type_membre']
        statut = form.cleaned_data['statut']
        
        # Préparer les résultats
        resultats = {
            'importes': 0,
            'erreurs': 0,
            'maj': 0,
            'messages': []
        }
        
        try:
            if extension == 'csv':
                self._process_csv(fichier, form, type_membre, statut, resultats)
            else:  # xlsx ou xls
                self._process_excel(fichier, type_membre, statut, resultats)
                
            # Message de succès
            message = _("Importation terminée: %(importes)s membres importés, %(maj)s mis à jour, %(erreurs)s erreurs.") % {
                'importes': resultats['importes'],
                'maj': resultats['maj'],
                'erreurs': resultats['erreurs']
            }
            messages.success(self.request, message)
            
            # Ajouter des messages détaillés si nécessaire
            for msg in resultats['messages']:
                messages.info(self.request, msg)
            
        except Exception as e:
            logger.error(f"Erreur lors de l'importation: {str(e)}")
            messages.error(self.request, _("Une erreur s'est produite lors de l'importation: %(error)s") % {'error': str(e)})
        
        return super().form_valid(form)
    
    def _process_csv(self, fichier, form, type_membre, statut, resultats):
        """Traiter un fichier CSV"""
        delimiter = form.cleaned_data['delimiter']
        has_header = form.cleaned_data['header']
        
        # Lire le fichier CSV
        decoded_file = fichier.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        
        reader = csv.reader(io_string, delimiter=delimiter)
        rows = list(reader)  # Convertir en liste pour pouvoir l'analyser avant transaction
        
        # Ignorer la première ligne si c'est un en-tête
        start_idx = 1 if has_header else 0
        
        # Vérification préliminaire des données avant de commencer la transaction
        fatal_error = False
        for i, row in enumerate(rows[start_idx:], start=start_idx+1):
            if len(row) < 3:  # Minimum: nom, prénom, email
                resultats['messages'].append(_("Ligne %(line)d: Données insuffisantes (minimum 3 colonnes requis)") % {'line': i})
                fatal_error = True
        
        if fatal_error:
            resultats['erreurs'] += 1
            raise ValueError(_("Le fichier CSV contient des erreurs structurelles. Veuillez corriger le format avant d'importer."))
        
        # Si tout est structurellement correct, commencer la transaction
        try:
            with transaction.atomic():
                for i, row in enumerate(rows[start_idx:], start=start_idx+1):
                    try:
                        # Récupérer les données de base
                        nom = row[0].strip() if row[0] else ""
                        prenom = row[1].strip() if row[1] else ""
                        email = row[2].strip() if row[2] else ""
                        
                        if not nom or not prenom or not email:
                            resultats['messages'].append(_("Ligne %(line)d: Données incomplètes (nom, prénom ou email manquant)") % {'line': i})
                            resultats['erreurs'] += 1
                            continue
                        
                        # Vérifier si le membre existe déjà
                        membre_existant = Membre.objects.filter(email=email).first()
                        
                        if membre_existant:
                            # Mettre à jour le membre existant si nécessaire
                            if membre_existant.nom != nom or membre_existant.prenom != prenom:
                                membre_existant.nom = nom
                                membre_existant.prenom = prenom
                                membre_existant.save(update_fields=['nom', 'prenom'])
                                resultats['maj'] += 1
                            
                            # Ajouter le type de membre s'il n'existe pas déjà
                            if type_membre and not membre_existant.est_type_actif(type_membre):
                                membre_existant.ajouter_type(type_membre)
                        else:
                            # Créer un nouveau membre
                            date_adhesion = timezone.now().date()
                            
                            # Données optionnelles (si disponibles)
                            telephone = row[3].strip() if len(row) > 3 and row[3] else None
                            adresse = row[4].strip() if len(row) > 4 and row[4] else None
                            
                            # Créer le membre
                            membre = Membre.objects.create(
                                nom=nom,
                                prenom=prenom,
                                email=email,
                                telephone=telephone,
                                adresse=adresse,
                                date_adhesion=date_adhesion,
                                statut=statut
                            )
                            
                            # Ajouter le type de membre si spécifié
                            if type_membre:
                                membre.ajouter_type(type_membre)
                            
                            resultats['importes'] += 1
                    
                    except Exception as e:
                        logger.error(f"Erreur ligne {i}: {str(e)}", exc_info=True)
                        resultats['messages'].append(_("Ligne %(line)d: %(error)s") % {'line': i, 'error': str(e)})
                        resultats['erreurs'] += 1
                        # Ne pas lever l'exception ici pour continuer à traiter les autres lignes
                    
                # Vérifier si le taux d'erreur est acceptable (par exemple, < 50%)
                if resultats['erreurs'] > 0 and resultats['erreurs'] / (resultats['importes'] + resultats['erreurs'] + resultats['maj']) > 0.5:
                    raise ValueError(_("Trop d'erreurs détectées lors de l'importation. L'opération a été annulée."))
                    
        except Exception as e:
            # Capture toute exception non gérée ou levée intentionnellement
            logger.error(f"Erreur lors de l'importation CSV: {str(e)}", exc_info=True)
            resultats['messages'].append(_("Erreur critique: %(error)s") % {'error': str(e)})
            resultats['erreurs'] += 1
            raise
    
    def _process_excel(self, fichier, type_membre, statut, resultats):
        """Traiter un fichier Excel"""
        wb = load_workbook(fichier)
        sheet = wb.active
        
        # Déterminer si la première ligne est un en-tête (on suppose que oui)
        has_header = True
        
        with transaction.atomic():
            rows = list(sheet.rows)
            start_idx = 1 if has_header else 0
            
            for i, row in enumerate(rows[start_idx:], start=start_idx+1):
                try:
                    # Récupérer les valeurs des cellules
                    nom = row[0].value.strip() if row[0].value else ""
                    prenom = row[1].value.strip() if row[1].value else ""
                    email = row[2].value.strip() if row[2].value else ""
                    
                    if not nom or not prenom or not email:
                        resultats['messages'].append(_("Ligne %(line)d: Données incomplètes") % {'line': i})
                        resultats['erreurs'] += 1
                        continue
                    
                    # Vérifier si le membre existe déjà
                    membre_existant = Membre.objects.filter(email=email).first()
                    
                    if membre_existant:
                        # Mettre à jour le membre existant si nécessaire
                        if membre_existant.nom != nom or membre_existant.prenom != prenom:
                            membre_existant.nom = nom
                            membre_existant.prenom = prenom
                            membre_existant.save(update_fields=['nom', 'prenom'])
                            resultats['maj'] += 1
                        
                        # Ajouter le type de membre s'il n'existe pas déjà
                        if not membre_existant.est_type_actif(type_membre):
                            membre_existant.ajouter_type(type_membre)
                    else:
                        # Créer un nouveau membre
                        date_adhesion = timezone.now().date()
                        
                        # Données optionnelles (si disponibles)
                        telephone = row[3].value if len(row) > 3 and row[3].value else None
                        adresse = row[4].value if len(row) > 4 and row[4].value else None
                        
                        # Créer le membre
                        membre = Membre.objects.create(
                            nom=nom,
                            prenom=prenom,
                            email=email,
                            telephone=telephone,
                            adresse=adresse,
                            date_adhesion=date_adhesion,
                            statut=statut
                        )
                        
                        # Ajouter le type de membre
                        membre.ajouter_type(type_membre)
                        
                        resultats['importes'] += 1
                
                except Exception as e:
                    logger.error(f"Erreur ligne {i}: {str(e)}")
                    resultats['messages'].append(_("Ligne %(line)d: %(error)s") % {'line': i, 'error': str(e)})
                    resultats['erreurs'] += 1


class MembreExportView(StaffRequiredMixin, View):
    """
    Vue pour exporter la liste des membres au format CSV ou Excel
    """
    # Correction pour MembreExportView.get() - Implémentation du filtre des cotisations impayées
    def get(self, request):
        format_export = request.GET.get('format', 'csv')
        
        # Récupérer les filtres de recherche pour les appliquer à l'export
        form = MembreSearchForm(request.GET)
        queryset = Membre.objects.all()
        
        if form.is_valid():
            # Appliquer les mêmes filtres que la vue liste
            if term := form.cleaned_data.get('terme'):
                queryset = queryset.recherche(term)
            
            if type_membre := form.cleaned_data.get('type_membre'):
                queryset = queryset.par_type(type_membre.id)
            
            if statut := form.cleaned_data.get('statut'):
                queryset = queryset.par_statut(statut.id)
            
            if date_min := form.cleaned_data.get('date_adhesion_min'):
                queryset = queryset.filter(date_adhesion__gte=date_min)
                
            if date_max := form.cleaned_data.get('date_adhesion_max'):
                queryset = queryset.filter(date_adhesion__lte=date_max)
            
            if age_min := form.cleaned_data.get('age_min'):
                queryset = queryset.par_age(age_min=age_min)
                
            if age_max := form.cleaned_data.get('age_max'):
                queryset = queryset.par_age(age_max=age_max)
            
            # Correction: Utiliser la même logique que dans MembreListView au lieu d'appeler une méthode inexistante
            if form.cleaned_data.get('cotisations_impayees'):
                try:
                    from apps.cotisations.models import Cotisation
                    from django.db.models import Count, Case, When, IntegerField
                    
                    # Cette requête annotée compte les cotisations impayées pour chaque membre
                    membres_avec_comptage = Cotisation.objects.values('membre_id').annotate(
                        nb_impayees=Count(
                            Case(
                                When(~Q(statut_paiement='payée'), then=1),
                                output_field=IntegerField()
                            )
                        )
                    ).filter(nb_impayees__gt=0)
                    
                    # Extraire uniquement les IDs des membres
                    membres_avec_impayees = [item['membre_id'] for item in membres_avec_comptage]
                    
                    if membres_avec_impayees:
                        # Filtrer le queryset pour ne garder que les membres avec cotisations impayées
                        queryset = queryset.filter(id__in=membres_avec_impayees)
                    else:
                        # Aucune cotisation impayée trouvée
                        queryset = queryset.none()  # Retourne un queryset vide
                except ImportError:
                    # Le module cotisations n'est pas disponible
                    pass
                except Exception:
                    # Capturer toute autre exception pour éviter l'échec silencieux
                    pass
            
            if compte := form.cleaned_data.get('avec_compte'):
                if compte == 'avec':
                    queryset = queryset.avec_compte_utilisateur()
                elif compte == 'sans':
                    queryset = queryset.sans_compte_utilisateur()
            
            if actif := form.cleaned_data.get('actif'):
                if actif == 'actif':
                    queryset = queryset.actifs()
                elif actif == 'inactif':
                    queryset = queryset.inactifs()
        
        # Précharger les relations pour optimiser
        membres = queryset.select_related('statut').prefetch_related('types')
        
        # Définir les champs à exporter
        champs = [
            'id', 'nom', 'prenom', 'email', 'telephone', 'adresse', 
            'code_postal', 'ville', 'pays', 'date_adhesion', 'date_naissance',
            'statut', 'types_actifs'
        ]
        
        # Générer le nom du fichier avec la date
        date_str = timezone.now().strftime('%Y%m%d')
        
        if format_export == 'excel':
            return self._export_excel(membres, champs, date_str)
        else:  # csv par défaut
            return self._export_csv(membres, champs, date_str)
    
    def _export_csv(self, membres, champs, date_str):
        """Exporter au format CSV avec encodage UTF-8 et BOM pour Excel"""
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="membres_{date_str}.csv"'
        
        # Ajouter un BOM (Byte Order Mark) pour que Excel reconnaisse correctement l'UTF-8
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')  # Utiliser point-virgule pour Excel
        
        # Écrire l'en-tête
        header = [
            _('ID'), _('Nom'), _('Prénom'), _('Email'), _('Téléphone'), 
            _('Adresse'), _('Code postal'), _('Ville'), _('Pays'),
            _('Date d\'adhésion'), _('Date de naissance'),
            _('Statut'), _('Types de membre')
        ]
        writer.writerow(header)
        
        # Écrire les données
        for membre in membres:
            row = []
            for champ in champs:
                if champ == 'statut':
                    row.append(membre.statut.nom if membre.statut else '')
                elif champ == 'types_actifs':
                    types = ", ".join([t.libelle for t in membre.get_types_actifs()])
                    row.append(types)
                elif champ in ('date_adhesion', 'date_naissance') and getattr(membre, champ):
                    # Formater les dates au format européen
                    date_value = getattr(membre, champ)
                    row.append(date_value.strftime('%d/%m/%Y'))
                else:
                    row.append(getattr(membre, champ, ''))
            writer.writerow(row)
        
        return response

    def _export_excel(self, membres, champs, date_str):
        """Exporter au format Excel avec prise en charge correcte des caractères accentués"""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="membres_{date_str}.xlsx"'
        
        # Créer un nouveau classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = _("Membres")
        
        # Écrire l'en-tête
        header = [
            _('ID'), _('Nom'), _('Prénom'), _('Email'), _('Téléphone'), 
            _('Adresse'), _('Code postal'), _('Ville'), _('Pays'),
            _('Date d\'adhésion'), _('Date de naissance'),
            _('Statut'), _('Types de membre')
        ]
        ws.append(header)
        
        # Écrire les données
        for membre in membres:
            row = []
            for champ in champs:
                if champ == 'statut':
                    row.append(membre.statut.nom if membre.statut else '')
                elif champ == 'types_actifs':
                    types = ", ".join([t.libelle for t in membre.get_types_actifs()])
                    row.append(types)
                elif champ in ('date_adhesion', 'date_naissance') and getattr(membre, champ):
                    # Pour Excel, on peut garder les objets date directement
                    row.append(getattr(membre, champ))
                else:
                    value = getattr(membre, champ, '')
                    # S'assurer que les valeurs None sont converties en chaînes vides
                    row.append(str(value) if value is not None else '')
            ws.append(row)
        
        # Style de l'en-tête
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        # Ajuster la largeur en fonction de la longueur de la valeur
                        # en tenant compte des caractères spéciaux
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            # Largeur minimale de 8 caractères
            adjusted_width = max(max_length + 2, 8)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Formater les dates
        date_format = 'DD/MM/YYYY'
        date_style = openpyxl.styles.NamedStyle(name='date_style', number_format=date_format)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            # Colonnes des dates (9 = date_adhesion, 10 = date_naissance)
            for col_idx in [9, 10]:
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, datetime.date):
                    cell.style = date_style
        
        # Enregistrer le classeur dans la réponse
        wb.save(response)
        
        return response


class MembreHistoriqueView(DetailView):
    """
    Vue pour afficher l'historique complet des modifications d'un membre
    """
    model = Membre
    template_name = 'membres/historique.html'
    context_object_name = 'membre'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Historique des modifications
        context['historique'] = HistoriqueMembre.objects.filter(
            membre=self.object
        ).select_related('utilisateur').order_by('-created_at')
        
        # Historique des types de membre
        context['historique_types'] = MembreTypeMembre.objects.filter(
            membre=self.object
        ).select_related('type_membre', 'modifie_par').order_by('-date_debut')
        
        return context


class MembreStatistiquesView(StaffRequiredMixin, TemplateView):
    """
    Vue pour afficher des statistiques détaillées sur les membres
    """
    template_name = 'membres/statistiques.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        # Statistiques générales
        total_membres = Membre.objects.count()
        membres_actifs = Membre.objects.actifs().count()
        
        # Répartition par type de membre
        types_stats = TypeMembre.objects.annotate(
            count=Count('membres_historique__membre', 
                          filter=Q(membres_historique__date_debut__lte=today, 
                                 membres_historique__date_fin__isnull=True),
                          distinct=True)
        ).order_by('-count')
        
        # Répartition par statut
        statuts_stats = Statut.objects.filter(
            membre__isnull=False
        ).annotate(
            count=Count('membre', distinct=True)
        ).order_by('-count')
        
        # Répartition par mois d'adhésion
        membres_par_mois = []
        for i in range(1, 13):
            month_name = datetime(2000, i, 1).strftime('%B')
            count = Membre.objects.filter(date_adhesion__month=i).count()
            membres_par_mois.append({'month': month_name, 'count': count})
        
        # Répartition par année d'adhésion
        annee_actuelle = today.year
        membres_par_annee = []
        for i in range(annee_actuelle - 9, annee_actuelle + 1):
            count = Membre.objects.filter(date_adhesion__year=i).count()
            membres_par_annee.append({'year': i, 'count': count})
        
        # Répartition par tranche d'âge
        membres_avec_age = Membre.objects.filter(date_naissance__isnull=False)
        tranches_age = [
            {'label': '< 18 ans', 'count': membres_avec_age.par_age(age_max=17).count()},
            {'label': '18-30 ans', 'count': membres_avec_age.par_age(age_min=18, age_max=30).count()},
            {'label': '31-45 ans', 'count': membres_avec_age.par_age(age_min=31, age_max=45).count()},
            {'label': '46-60 ans', 'count': membres_avec_age.par_age(age_min=46, age_max=60).count()},
            {'label': '61-75 ans', 'count': membres_avec_age.par_age(age_min=61, age_max=75).count()},
            {'label': '> 75 ans', 'count': membres_avec_age.par_age(age_min=76).count()},
        ]
        
        # Adhésions par année
        adhesions_par_annee = []
        for i in range(annee_actuelle - 9, annee_actuelle + 1):
            count = Membre.objects.filter(date_adhesion__year=i).count()
            adhesions_par_annee.append({'year': i, 'count': count})
        
        # Membres avec/sans compte utilisateur
        with_account = Membre.objects.filter(utilisateur__isnull=False).count()
        without_account = total_membres - with_account
        
        # Mettre à jour le contexte
        context.update({
            'total_membres': total_membres,
            'membres_actifs': membres_actifs,
            'pourcentage_actifs': round(membres_actifs / total_membres * 100 if total_membres else 0, 1),
            'types_stats': types_stats,
            'statuts_stats': statuts_stats,
            'membres_par_mois': membres_par_mois,
            'membres_par_annee': membres_par_annee,
            'tranches_age': tranches_age,
            'adhesions_par_annee': adhesions_par_annee,
            'with_account': with_account,
            'without_account': without_account,
            # Convertir en JSON pour les graphiques JavaScript
            'chart_types': json.dumps([{'name': t.libelle, 'value': t.count} for t in types_stats]),
            'chart_statuts': json.dumps([{'name': s.nom, 'value': s.count} for s in statuts_stats]),
            'chart_mois': json.dumps(membres_par_mois),
            'chart_annees': json.dumps(membres_par_annee),
            'chart_ages': json.dumps(tranches_age),
            'chart_adhesions': json.dumps(adhesions_par_annee),
            'chart_comptes': json.dumps([
                {'name': _('Avec compte'), 'value': with_account},
                {'name': _('Sans compte'), 'value': without_account}
            ]),
        })
        
        return context

class TypeMembreDetailView(StaffRequiredMixin, DetailView):
    """
    Vue pour afficher les détails d'un type de membre
    """
    model = TypeMembre
    template_name = 'membres/type_membre_detail.html'
    context_object_name = 'type_membre'
    
    def get_object(self, queryset=None):
        """
        Récupérer l'objet sans tenter d'annoter ou de compter les membres,
        ce qui peut causer des erreurs avec les relations.
        """
        if queryset is None:
            queryset = self.get_queryset()
        
        # Utiliser la méthode standard pour récupérer l'objet par pk
        pk = self.kwargs.get(self.pk_url_kwarg)
        queryset = queryset.filter(pk=pk)
        
        try:
            # Get the single item from the filtered queryset
            obj = queryset.get()
        except queryset.model.DoesNotExist:
            raise Http404(_("Aucun type de membre trouvé avec cet identifiant"))
            
        # Ajouter dynamiquement une méthode nb_membres_actifs à l'objet
        type_membre = obj
        today = timezone.now().date()
        
        # Calculer le nombre de membres actifs
        membres_count = type_membre.membres_historique.filter(
            date_debut__lte=today,
            date_fin__isnull=True
        ).values('membre').distinct().count()
        
        # Ajouter une méthode dynamique à l'objet qui renvoie ce nombre
        def nb_membres_actifs(self):
            return membres_count
            
        # Lier la méthode à l'instance
        type_membre.nb_membres_actifs = types.MethodType(nb_membres_actifs, type_membre)
        
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ajouter les membres actifs au contexte en utilisant la relation correcte
        today = timezone.now().date()
        context['membres_actifs'] = self.object.membres_historique.filter(
            date_debut__lte=today,
            date_fin__isnull=True
        ).values('membre').distinct()
        return context
    
class MembreCorbeillePage(StaffRequiredMixin, TrashViewMixin, ListView):
    """
    Vue pour afficher les membres supprimés logiquement
    """
    model = Membre
    template_name = 'membres/corbeille.html'
    context_object_name = 'membres'
    paginate_by = 20
    
    def get_queryset(self):
        # N'afficher que les membres supprimés
        return Membre.objects.only_deleted()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _("Corbeille - Membres supprimés")
        context['total_supprimes'] = Membre.objects.only_deleted().count()
        return context
    
class MembreRestaurerView(StaffRequiredMixin, View):
    """
    Vue pour restaurer un membre supprimé
    """
    def post(self, request, pk):
        membre = get_object_or_404(Membre.objects.only_deleted(), pk=pk)
        membre.deleted_at = None
        membre.save(update_fields=['deleted_at'])
        
        # Ajouter à l'historique
        HistoriqueMembre.objects.create(
            membre=membre,
            utilisateur=request.user,
            action='restauration',
            description=_("Restauration du membre"),
            donnees_avant={'deleted_at': str(membre.deleted_at)},
            donnees_apres={'deleted_at': None}
        )
        
        messages.success(
            request, 
            _("Le membre %(name)s a été restauré avec succès.") % {'name': membre.nom_complet}
        )
        return redirect('membres:membre_detail', pk=membre.pk)
    
class MembreRestaurerView(RestoreViewMixin, View):
    """Vue pour restaurer un membre depuis la corbeille."""
    model = Membre
    success_url = reverse_lazy('membres:membre_liste')
    
    def get_object(self):
        """Récupérer le membre à restaurer."""
        return get_object_or_404(Membre.objects.only_deleted(), pk=self.kwargs['pk'])
    
    def get_success_response(self):
        """Response après restauration réussie."""
        messages.success(
            self.request, 
            _("Le membre a été restauré avec succès.")
        )
        return redirect(self.success_url)

class MembreSuppressionDefinitiveView(RestoreViewMixin, View):
    """Vue pour supprimer définitivement un membre."""
    model = Membre
    success_url = reverse_lazy('membres:membre_corbeille')
    
    def get_object(self):
        """Récupérer le membre à supprimer définitivement."""
        return get_object_or_404(Membre.objects.only_deleted(), pk=self.kwargs['pk'])
    
    def post(self, request, *args, **kwargs):
        """Supprimer définitivement le membre."""
        membre = self.get_object()
        
        # Journaliser l'action avant suppression définitive
        HistoriqueMembre.objects.create(
            membre=membre,
            utilisateur=request.user,
            action='suppression_definitive',
            description=_("Suppression définitive du membre"),
            donnees_avant={
                'nom': membre.nom,
                'prenom': membre.prenom,
                'email': membre.email,
                'deleted_at': str(membre.deleted_at)
            }
        )
        
        # Suppression physique
        membre.delete(hard=True)
        
        messages.success(
            request, 
            _("Le membre a été définitivement supprimé.")
        )
        return redirect(self.success_url)

class MembreUpdateView(StaffRequiredMixin, UpdateView):
    """
    Vue pour modifier un membre existant
    """
    model = Membre
    form_class = MembreForm
    template_name = 'membres/form.html'
    success_url = reverse_lazy('membres:membre_liste')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        membre = form.save()
        messages.success(
            self.request, 
            _("Le membre %(name)s a été modifié avec succès.") % {'name': membre.nom_complet}
        )
        return super().form_valid(form)
    
class GuideIntegrationView(StaffRequiredMixin, TemplateView):
    """
    Vue pour afficher le guide d'intégration pour les nouveaux membres
    """
    template_name = 'membres/guide_integration.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Récupérer le membre associé à l'utilisateur connecté
        try:
            context['membre'] = Membre.objects.get(utilisateur=self.request.user)
        except Membre.DoesNotExist:
            context['membre'] = None
            
        # Informations supplémentaires pour personnaliser le guide
        if 'pk' in kwargs:
            try:
                membre = Membre.objects.get(pk=kwargs['pk'])
                # Vérifier que l'utilisateur est administrateur ou le propriétaire du profil
                if self.request.user.is_staff or (context['membre'] and context['membre'].pk == membre.pk):
                    context['membre'] = membre
            except Membre.DoesNotExist:
                pass
        
        return context
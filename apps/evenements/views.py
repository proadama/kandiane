# apps/evenements/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView,
    FormView, TemplateView, View
)
from django.http import JsonResponse, HttpResponse, Http404
from django.utils import timezone
from django.db.models import Q, Count, Sum, Avg
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.core.paginator import Paginator
from django.db import transaction
from datetime import datetime, timedelta
import json

from django.contrib.syndication.views import Feed
from django.utils.feedgenerator import Atom1Feed
from django.core.management import call_command
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.core.cache import cache
from django.contrib.contenttypes.models import ContentType
import openpyxl
from io import BytesIO
import xml.etree.ElementTree as ET

from apps.core.mixins import StaffRequiredMixin, PermissionRequiredMixin, AjaxRequiredMixin
from apps.membres.models import Membre
from .models import (
    Evenement, TypeEvenement, InscriptionEvenement, 
    AccompagnantInvite, ValidationEvenement, SessionEvenement,
    EvenementRecurrence
)
from .forms import (
    EvenementForm, InscriptionEvenementForm, EvenementSearchForm,
    ValidationEvenementForm, AccompagnantForm, EvenementRecurrenceForm,
    SessionEvenementForm, PaiementInscriptionForm, TypeEvenementForm,
    ExportEvenementsForm
)


class DashboardEvenementView(LoginRequiredMixin, TemplateView):
    """
    Tableau de bord accessible à tous les utilisateurs connectés
    """
    template_name = 'evenements/rapports/dashboard.html'  # CORRECTION du template
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Données pour tous les utilisateurs connectés
        context['evenements_publics'] = Evenement.objects.filter(
            statut='publie',
            date_debut__gte=timezone.now()
        ).order_by('date_debut')[:5]
        
        # Données supplémentaires pour les membres
        if hasattr(user, 'membre_utilisateur'):
            try:
                membre = user.membre_utilisateur
                context.update({
                    'mes_prochaines_inscriptions': InscriptionEvenement.objects.filter(
                        membre=membre,
                        evenement__date_debut__gte=timezone.now(),
                        statut__in=['confirmee', 'en_attente']
                    ).select_related('evenement')[:5],
                    'historique_participations': InscriptionEvenement.objects.filter(
                        membre=membre, statut='presente'
                    ).count(),
                })
            except:
                pass
        
        # Données administratives pour staff uniquement
        if user.is_staff:
            context.update({
                'total_evenements': Evenement.objects.count(),
                'evenements_publies': Evenement.objects.filter(statut='publie').count(),
                'evenements_a_valider': ValidationEvenement.objects.filter(
                    statut_validation='en_attente'
                ).count(),
                'inscriptions_en_attente': InscriptionEvenement.objects.filter(
                    statut='en_attente'
                ).count(),
                'prochains_evenements': Evenement.objects.filter(
                    date_debut__gte=timezone.now()
                ).order_by('date_debut')[:5],
            })
        
        return context
    
# =============================================================================
# VUES ÉVÉNEMENTS
# =============================================================================

class EvenementListView(LoginRequiredMixin, ListView):
    """
    Liste des événements accessible aux membres connectés
    """
    model = Evenement
    template_name = 'evenements/liste.html'
    context_object_name = 'evenements'
    paginate_by = 20
    
    # CORRECTION : Pas de StaffRequiredMixin ici
    def get_queryset(self):
        # Tous les utilisateurs connectés voient les événements publiés
        if self.request.user.is_staff:
            queryset = Evenement.objects.all()
        else:
            queryset = Evenement.objects.filter(statut='publie')
        
        # Appliquer les filtres de recherche
        form = EvenementSearchForm(self.request.GET)
        if form.is_valid():
            if form.cleaned_data.get('titre'):
                queryset = queryset.filter(titre__icontains=form.cleaned_data['titre'])
            if form.cleaned_data.get('type_evenement'):
                queryset = queryset.filter(type_evenement=form.cleaned_data['type_evenement'])
            if form.cleaned_data.get('date_debut'):
                queryset = queryset.filter(date_debut__gte=form.cleaned_data['date_debut'])
        
        return queryset.select_related('type_evenement', 'organisateur')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = EvenementSearchForm(self.request.GET)
        
        # CORRECTION : Utiliser les URLs directes au lieu des namespaces
        try:
            context['ajax_urls'] = {
                'autocomplete_organisateurs': reverse('evenements:autocomplete_organisateurs'),
                'autocomplete_lieux': reverse('evenements:autocomplete_lieux'),
            }
        except:
            # Fallback si les URLs n'existent pas encore
            context['ajax_urls'] = {
                'autocomplete_organisateurs': '#',
                'autocomplete_lieux': '#',
            }
        
        return context

class EvenementDetailView(LoginRequiredMixin, DetailView):
    """
    Détail d'un événement
    """
    model = Evenement
    template_name = 'evenements/detail.html'
    context_object_name = 'evenement'
    
    def get_queryset(self):
        queryset = Evenement.objects.select_related('type_evenement', 'organisateur')
        
        # Filtrage selon les permissions
        if not self.request.user.is_staff:
            queryset = queryset.filter(statut='publie')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        evenement = self.get_object()
        user = self.request.user
        
        context['ajax_urls'] = {
            'places_disponibles': f'/evenements/ajax/evenements/{evenement.pk}/places-disponibles/',
            'calculer_tarif': f'/evenements/ajax/evenements/{evenement.pk}/calculer-tarif/',
        }
        
        # Sessions de l'événement
        context['sessions'] = evenement.sessions.all().order_by('ordre_session')
        
        # Vérifier si l'utilisateur peut s'inscrire
        try:
            membre = Membre.objects.get(utilisateur=user)
            peut_inscrire, message = evenement.peut_s_inscrire(membre)
            context['peut_s_inscrire'] = peut_inscrire
            context['message_inscription'] = message
            
            # Vérifier si déjà inscrit
            try:
                inscription_existante = InscriptionEvenement.objects.get(
                    evenement=evenement, membre=membre
                )
                context['inscription_existante'] = inscription_existante
            except InscriptionEvenement.DoesNotExist:
                context['inscription_existante'] = None
                
        except Membre.DoesNotExist:
            context['peut_s_inscrire'] = False
            context['message_inscription'] = "Vous devez être membre pour vous inscrire"
        
        # Permissions d'édition
        context['peut_modifier'] = (
            user.is_staff or 
            (hasattr(user, 'membre') and evenement.organisateur == user)
        )
        
        return context


class EvenementCreateView(StaffRequiredMixin, CreateView):
    """
    Création d'un nouvel événement
    """
    model = Evenement
    form_class = EvenementForm
    template_name = 'evenements/form.html'
    success_url = reverse_lazy('evenements:liste')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # CORRECTION : Redirection directe au lieu du namespace
        if self.object.statut == 'en_attente_validation':
            messages.success(
                self.request, 
                f"L'événement '{self.object.titre}' a été créé et envoyé pour validation."
            )
            # Rediriger vers la liste au lieu de validation
            return redirect('evenements:liste')
        else:
            messages.success(
                self.request, 
                f"L'événement '{self.object.titre}' a été créé et publié."
            )
        
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titre_page'] = 'Créer un événement'
        context['action'] = 'Créer'
        return context


class EvenementUpdateView(LoginRequiredMixin, UpdateView):
    """
    Modification d'un événement
    """
    model = Evenement
    form_class = EvenementForm
    template_name = 'evenements/form.html'
    
    def get_queryset(self):
        queryset = Evenement.objects.all()
        
        # Limiter aux événements que l'utilisateur peut modifier
        if not self.request.user.is_staff:
            queryset = queryset.filter(organisateur=self.request.user)
        
        return queryset
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f"L'événement '{self.object.titre}' a été modifié."
        )
        return response
    
    def get_success_url(self):
        return reverse('evenements:detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['titre_page'] = f'Modifier {self.object.titre}'
        context['action'] = 'Modifier'
        return context


class EvenementDeleteView(StaffRequiredMixin, DeleteView):
    """
    Suppression (logique) d'un événement
    """
    model = Evenement
    template_name = 'evenements/confirmer_suppression.html'
    success_url = reverse_lazy('evenements:liste')
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Vérifier s'il y a des inscriptions confirmées
        inscriptions_confirmees = self.object.inscriptions.confirmees().count()
        if inscriptions_confirmees > 0:
            messages.error(
                request,
                f"Impossible de supprimer l'événement : {inscriptions_confirmees} inscription(s) confirmée(s)."
            )
            return redirect('evenements:detail', pk=self.object.pk)
        
        # Suppression logique
        self.object.delete()
        messages.success(
            request,
            f"L'événement '{self.object.titre}' a été supprimé."
        )
        
        return redirect(self.success_url)


# =============================================================================
# VUES INSCRIPTIONS
# =============================================================================

class InscriptionCreateView(LoginRequiredMixin, CreateView):
    """
    Inscription à un événement
    """
    model = InscriptionEvenement
    form_class = InscriptionEvenementForm
    template_name = 'evenements/inscription/form.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.evenement = get_object_or_404(Evenement, pk=kwargs['evenement_pk'])
        
        # Vérifier que l'utilisateur est un membre
        try:
            self.membre = Membre.objects.get(utilisateur=request.user)
        except Membre.DoesNotExist:
            messages.error(request, "Vous devez être membre pour vous inscrire à un événement.")
            return redirect('evenements:detail', pk=self.evenement.pk)
        
        # Vérifier si peut s'inscrire
        peut_inscrire, message = self.evenement.peut_s_inscrire(self.membre)
        if not peut_inscrire:
            messages.error(request, message)
            return redirect('evenements:detail', pk=self.evenement.pk)
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['evenement'] = self.evenement
        kwargs['membre'] = self.membre
        return kwargs
    
    def form_valid(self, form):
        with transaction.atomic():
            response = super().form_valid(form)
            
            # Message selon le statut
            if self.object.statut == 'en_attente':
                messages.success(
                    self.request,
                    f"Votre inscription à '{self.evenement.titre}' a été enregistrée. "
                    f"Vous avez {self.evenement.delai_confirmation}h pour la confirmer."
                )
            elif self.object.statut == 'liste_attente':
                messages.info(
                    self.request,
                    f"L'événement '{self.evenement.titre}' est complet. "
                    "Vous avez été placé(e) en liste d'attente."
                )
            
            return response
    
    def get_success_url(self):
        # CORRECTION : Utiliser l'URL correcte
        return reverse('evenements:inscription_detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evenement'] = self.evenement
        context['membre'] = self.membre
        context['tarif_applicable'] = self.evenement.calculer_tarif_membre(self.membre)
        return context


class InscriptionDetailView(LoginRequiredMixin, DetailView):
    """
    Détail d'une inscription
    """
    model = InscriptionEvenement
    template_name = 'evenements/inscription/detail.html'
    context_object_name = 'inscription'
    
    def get_queryset(self):
        queryset = InscriptionEvenement.objects.select_related(
            'evenement', 'membre', 'mode_paiement'
        ).prefetch_related('accompagnants')
        
        # Limiter aux inscriptions de l'utilisateur ou staff
        if not self.request.user.is_staff:
            try:
                membre = Membre.objects.get(utilisateur=self.request.user)
                queryset = queryset.filter(membre=membre)
            except Membre.DoesNotExist:
                queryset = queryset.none()
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inscription = self.get_object()
        
        context['accompagnants'] = inscription.accompagnants.all()
        context['montant_total'] = inscription.calculer_montant_total()
        context['montant_restant'] = inscription.montant_restant
        context['peut_confirmer'] = inscription.statut == 'en_attente'
        context['peut_annuler'] = inscription.statut in ['en_attente', 'confirmee']
        
        return context


class ConfirmerInscriptionView(LoginRequiredMixin, View):
    """
    Confirmation d'une inscription
    """
    
    def post(self, request, pk):
        inscription = get_object_or_404(InscriptionEvenement, pk=pk)
        
        # Vérifier les permissions
        if not request.user.is_staff and inscription.membre.utilisateur != request.user:
            messages.error(request, "Vous n'avez pas l'autorisation de confirmer cette inscription.")
            return redirect('evenements:dashboard')
        
        # Confirmer l'inscription
        if inscription.confirmer_inscription():
            messages.success(request, "Votre inscription a été confirmée.")
        else:
            messages.error(request, "Impossible de confirmer cette inscription.")
        
        # CORRECTION : Utiliser l'URL correcte
        return redirect('evenements:inscription_detail', pk=inscription.pk)


class AnnulerInscriptionView(LoginRequiredMixin, View):
    """
    Annulation d'une inscription
    """
    
    def post(self, request, pk):
        inscription = get_object_or_404(InscriptionEvenement, pk=pk)
        
        # Vérifier les permissions
        if not request.user.is_staff and inscription.membre.utilisateur != request.user:
            messages.error(request, "Vous n'avez pas l'autorisation d'annuler cette inscription.")
            return redirect('evenements:dashboard')
        
        # Annuler l'inscription
        raison = request.POST.get('raison', 'Annulation par l\'utilisateur')
        if inscription.annuler_inscription(raison):
            messages.success(request, "Votre inscription a été annulée.")
        else:
            messages.error(request, "Impossible d'annuler cette inscription.")
        
        return redirect('evenements:inscription_detail', pk=inscription.pk)


class ConfirmerInscriptionEmailView(View):
    """
    Confirmation d'inscription via lien email (accès public)
    """
    
    def get(self, request, code):
        try:
            inscription = InscriptionEvenement.objects.get(code_confirmation=code)
        except InscriptionEvenement.DoesNotExist:
            return render(request, 'evenements/inscription/confirmation_erreur.html', {
                'erreur': 'Code de confirmation invalide.'
            })
        
        # Vérifier si la confirmation est encore possible
        if inscription.statut != 'en_attente':
            return render(request, 'evenements/inscription/confirmation_erreur.html', {
                'erreur': 'Cette inscription a déjà été traitée.'
            })
        
        if inscription.est_en_retard_confirmation:
            return render(request, 'evenements/inscription/confirmation_erreur.html', {
                'erreur': 'Le délai de confirmation a expiré.'
            })
        
        return render(request, 'evenements/inscription/confirmation.html', {
            'inscription': inscription
        })
    
    def post(self, request, code):
        try:
            inscription = InscriptionEvenement.objects.get(code_confirmation=code)
        except InscriptionEvenement.DoesNotExist:
            return render(request, 'evenements/inscription/confirmation_erreur.html', {
                'erreur': 'Code de confirmation invalide.'
            })
        
        action = request.POST.get('action')
        
        if action == 'confirmer':
            if inscription.confirmer_inscription():
                return render(request, 'evenements/inscription/confirmation_succes.html', {
                    'inscription': inscription,
                    'message': 'Votre inscription a été confirmée avec succès.'
                })
        elif action == 'refuser':
            if inscription.annuler_inscription('Refus par email'):
                return render(request, 'evenements/inscription/confirmation_succes.html', {
                    'inscription': inscription,
                    'message': 'Votre inscription a été annulée.'
                })
        
        return render(request, 'evenements/inscription/confirmation_erreur.html', {
            'erreur': 'Une erreur est survenue lors du traitement de votre demande.'
        })


# =============================================================================
# VUES VALIDATION
# =============================================================================

class ValidationListView(StaffRequiredMixin, ListView):
    """
    Liste des événements à valider
    """
    model = ValidationEvenement
    template_name = 'evenements/validation/liste.html'
    context_object_name = 'validations'
    paginate_by = 20
    
    def get_queryset(self):
        return ValidationEvenement.objects.en_attente().select_related(
            'evenement', 'evenement__type_evenement', 'evenement__organisateur'
        ).order_by('evenement__date_debut')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['validations_urgentes'] = ValidationEvenement.objects.urgentes()
        return context


class ValidationDetailView(StaffRequiredMixin, DetailView):
    """
    Détail d'une validation d'événement
    """
    model = ValidationEvenement
    template_name = 'evenements/validation/detail.html'
    context_object_name = 'validation'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ValidationEvenementForm()
        return context


class ApprouverEvenementView(StaffRequiredMixin, View):
    """
    Approbation d'un événement
    """
    
    def post(self, request, pk):
        validation = get_object_or_404(ValidationEvenement, pk=pk)
        commentaire = request.POST.get('commentaire', '')
        
        validation.approuver(request.user, commentaire)
        messages.success(
            request, 
            f"L'événement '{validation.evenement.titre}' a été approuvé."
        )
        
        return redirect('evenements:validation_liste')


class RefuserEvenementView(StaffRequiredMixin, View):
    """
    Refus d'un événement
    """
    
    def post(self, request, pk):
        validation = get_object_or_404(ValidationEvenement, pk=pk)
        commentaire = request.POST.get('commentaire', '')
        
        if not commentaire:
            messages.error(request, "Un commentaire est obligatoire pour refuser un événement.")
            return redirect('evenements:validation_detail', pk=pk)
        
        validation.refuser(request.user, commentaire)
        messages.success(
            request, 
            f"L'événement '{validation.evenement.titre}' a été refusé."
        )
        
        return redirect('evenements:validation_liste')


# =============================================================================
# VUES UTILITAIRES ET AJAX
# =============================================================================

class CheckPlacesDisponiblesView(AjaxRequiredMixin, View):
    """
    Vérification AJAX des places disponibles
    """
    
    def get(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        
        return JsonResponse({
            'places_disponibles': evenement.places_disponibles,
            'est_complet': evenement.est_complet,
            'taux_occupation': evenement.taux_occupation,
        })


class MesInscriptionsView(LoginRequiredMixin, ListView):
    """
    Liste des inscriptions de l'utilisateur connecté
    """
    model = InscriptionEvenement
    template_name = 'evenements/inscription/mes_inscriptions.html'
    context_object_name = 'inscriptions'
    paginate_by = 20
    
    def get_queryset(self):
        try:
            membre = Membre.objects.get(utilisateur=self.request.user)
            return InscriptionEvenement.objects.par_membre(membre).select_related(
                'evenement', 'evenement__type_evenement'
            ).order_by('-date_inscription')
        except Membre.DoesNotExist:
            return InscriptionEvenement.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistiques personnelles
        if hasattr(self, 'membre'):
            stats = InscriptionEvenement.objects.statistiques_membre(self.membre)
            context['stats'] = stats
        
        return context


# =============================================================================
# VUES DE GESTION DES TYPES D'ÉVÉNEMENTS
# =============================================================================

class TypeEvenementListView(StaffRequiredMixin, ListView):
    """
    Liste des types d'événements
    """
    model = TypeEvenement
    template_name = 'evenements/types/liste.html'
    context_object_name = 'types'
    
    def get_queryset(self):
        return TypeEvenement.objects.par_ordre_affichage()


class TypeEvenementCreateView(StaffRequiredMixin, CreateView):
    """
    Création d'un type d'événement
    """
    model = TypeEvenement
    # CORRECTION : Utiliser un template simple ou créer le form.html basique
    template_name = 'evenements/form.html'  # au lieu de 'evenements/types/form.html'
    fields = ['libelle', 'description', 'couleur_affichage', 'permet_accompagnants', 'necessite_validation']
    success_url = reverse_lazy('evenements:types_liste')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f"Le type d'événement '{self.object.libelle}' a été créé."
        )
        return response


# =============================================================================
# PHASE 5.4.2 : VUES SPÉCIALISÉES
# =============================================================================

class CalendrierEvenementView(LoginRequiredMixin, TemplateView):
    """
    Vue calendrier interactive des événements
    """
    template_name = 'evenements/calendrier.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Récupérer les événements pour le calendrier
        evenements = Evenement.objects.publies().select_related('type_evenement')
        
        # Filtrer par mois si spécifié
        mois = self.request.GET.get('mois')
        annee = self.request.GET.get('annee')
        if mois and annee:
            try:
                date_debut = datetime(int(annee), int(mois), 1)
                if int(mois) == 12:
                    date_fin = datetime(int(annee) + 1, 1, 1) - timedelta(days=1)
                else:
                    date_fin = datetime(int(annee), int(mois) + 1, 1) - timedelta(days=1)
                
                evenements = evenements.filter(
                    date_debut__date__range=[date_debut.date(), date_fin.date()]
                )
            except (ValueError, TypeError):
                pass
        
        # Convertir en format JSON pour le calendrier
        evenements_json = []
        for evenement in evenements:
            evenements_json.append({
                'id': evenement.id,
                'title': evenement.titre,
                'start': evenement.date_debut.isoformat(),
                'end': evenement.date_fin.isoformat() if evenement.date_fin else None,
                'url': reverse('evenements:detail', kwargs={'pk': evenement.pk}),
                'backgroundColor': evenement.type_evenement.couleur_affichage,
                'borderColor': evenement.type_evenement.couleur_affichage,
                'extendedProps': {
                    'type': evenement.type_evenement.libelle,
                    'lieu': evenement.lieu,
                    'places_disponibles': evenement.places_disponibles,
                    'est_complet': evenement.est_complet,
                }
            })
        
        context['evenements_json'] = json.dumps(evenements_json)
        context['types_evenements'] = TypeEvenement.objects.all()
        
        return context


class ExportInscritsView(StaffRequiredMixin, View):
    """
    Export de la liste des inscrits à un événement
    """
    
    def get(self, request, evenement_pk):
        evenement = get_object_or_404(Evenement, pk=evenement_pk)
        format_export = request.GET.get('format', 'csv')
        
        # Récupérer les inscriptions confirmées
        inscriptions = InscriptionEvenement.objects.filter(
            evenement=evenement,
            statut__in=['confirmee', 'presente']
        ).select_related('membre').prefetch_related('accompagnants')
        
        if format_export == 'pdf':
            return self._export_pdf(evenement, inscriptions)
        elif format_export == 'excel':
            return self._export_excel(evenement, inscriptions)
        else:
            return self._export_csv(evenement, inscriptions)
    
    def _export_csv(self, evenement, inscriptions):
        """Export CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="inscrits_{evenement.reference}.csv"'
        
        # BOM pour Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow([
            'Nom', 'Prénom', 'Email', 'Téléphone', 'Date inscription', 
            'Statut', 'Nb accompagnants', 'Montant payé', 'Mode paiement'
        ])
        
        for inscription in inscriptions:
            writer.writerow([
                inscription.membre.nom,
                inscription.membre.prenom,
                inscription.membre.email,
                inscription.membre.telephone or '',
                inscription.date_inscription.strftime('%d/%m/%Y %H:%M'),
                inscription.get_statut_display(),
                inscription.nombre_accompagnants,
                f"{inscription.montant_paye}€",
                inscription.mode_paiement.libelle if inscription.mode_paiement else ''
            ])
            
            # Ajouter les accompagnants
            for accompagnant in inscription.accompagnants.all():
                writer.writerow([
                    f"  → {accompagnant.nom}",
                    accompagnant.prenom,
                    accompagnant.email,
                    accompagnant.telephone or '',
                    '',
                    'Accompagnant',
                    '',
                    '',
                    ''
                ])
        
        return response
    
    def _export_excel(self, evenement, inscriptions):
        """Export Excel"""
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inscrits"
        
        # En-têtes
        headers = [
            'Nom', 'Prénom', 'Email', 'Téléphone', 'Date inscription',
            'Statut', 'Nb accompagnants', 'Montant payé', 'Mode paiement'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        row = 2
        for inscription in inscriptions:
            ws.cell(row=row, column=1, value=inscription.membre.nom)
            ws.cell(row=row, column=2, value=inscription.membre.prenom)
            ws.cell(row=row, column=3, value=inscription.membre.email)
            ws.cell(row=row, column=4, value=inscription.membre.telephone or '')
            ws.cell(row=row, column=5, value=inscription.date_inscription.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=6, value=inscription.get_statut_display())
            ws.cell(row=row, column=7, value=inscription.nombre_accompagnants)
            ws.cell(row=row, column=8, value=f"{inscription.montant_paye}€")
            ws.cell(row=row, column=9, value=inscription.mode_paiement.libelle if inscription.mode_paiement else '')
            row += 1
            
            # Accompagnants
            for accompagnant in inscription.accompagnants.all():
                ws.cell(row=row, column=1, value=f"  → {accompagnant.nom}")
                ws.cell(row=row, column=2, value=accompagnant.prenom)
                ws.cell(row=row, column=3, value=accompagnant.email)
                ws.cell(row=row, column=4, value=accompagnant.telephone or '')
                ws.cell(row=row, column=6, value='Accompagnant')
                row += 1
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inscrits_{evenement.reference}.xlsx"'
        
        return response
    
    def _export_pdf(self, evenement, inscriptions):
        """Export PDF"""
        from django.http import HttpResponse
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from io import BytesIO
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Titre
        p.setFont("Helvetica-Bold", 16)
        p.drawString(2*cm, height - 2*cm, f"Liste des inscrits - {evenement.titre}")
        
        # Informations événement
        p.setFont("Helvetica", 10)
        y = height - 3*cm
        p.drawString(2*cm, y, f"Date: {evenement.date_debut.strftime('%d/%m/%Y à %H:%M')}")
        y -= 0.5*cm
        p.drawString(2*cm, y, f"Lieu: {evenement.lieu}")
        y -= 0.5*cm
        p.drawString(2*cm, y, f"Organisateur: {evenement.organisateur.get_full_name()}")
        
        # Liste des inscrits
        y -= 1.5*cm
        p.setFont("Helvetica-Bold", 12)
        p.drawString(2*cm, y, "Participants inscrits:")
        
        y -= 0.8*cm
        p.setFont("Helvetica", 9)
        
        for i, inscription in enumerate(inscriptions, 1):
            if y < 3*cm:  # Nouvelle page si nécessaire
                p.showPage()
                y = height - 2*cm
            
            text = f"{i}. {inscription.membre.prenom} {inscription.membre.nom}"
            if inscription.nombre_accompagnants > 0:
                text += f" (+{inscription.nombre_accompagnants} accompagnant(s))"
            
            p.drawString(2*cm, y, text)
            y -= 0.4*cm
            
            # Accompagnants
            for accompagnant in inscription.accompagnants.all():
                if y < 3*cm:
                    p.showPage()
                    y = height - 2*cm
                p.drawString(3*cm, y, f"• {accompagnant.prenom} {accompagnant.nom}")
                y -= 0.4*cm
        
        # Résumé
        y -= 1*cm
        p.setFont("Helvetica-Bold", 10)
        total_participants = sum(1 + i.nombre_accompagnants for i in inscriptions)
        p.drawString(2*cm, y, f"Total: {len(inscriptions)} inscriptions, {total_participants} participants")
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="inscrits_{evenement.reference}.pdf"'
        
        return response


class EvenementSearchView(LoginRequiredMixin, ListView):
    """
    Vue de recherche avancée d'événements
    """
    model = Evenement
    template_name = 'evenements/recherche.html'
    context_object_name = 'evenements'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Evenement.objects.publies().avec_statistiques()
        
        # Application des filtres avancés
        query = self.request.GET.get('q', '')
        if query:
            queryset = queryset.recherche(query)
        
        # Filtres par facettes
        type_ids = self.request.GET.getlist('types')
        if type_ids:
            queryset = queryset.filter(type_evenement__id__in=type_ids)
        
        # Filtres par prix
        prix_min = self.request.GET.get('prix_min')
        prix_max = self.request.GET.get('prix_max')
        if prix_min:
            queryset = queryset.filter(tarif_membre__gte=prix_min)
        if prix_max:
            queryset = queryset.filter(tarif_membre__lte=prix_max)
        
        # Tri
        sort = self.request.GET.get('sort', 'date_debut')
        if sort == 'titre':
            queryset = queryset.order_by('titre')
        elif sort == 'prix':
            queryset = queryset.order_by('tarif_membre')
        elif sort == 'popularite':
            queryset = queryset.order_by('-inscriptions_confirmees')
        else:
            queryset = queryset.order_by('date_debut')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['types_evenements'] = TypeEvenement.objects.all()
        context['query'] = self.request.GET.get('q', '')
        context['selected_types'] = self.request.GET.getlist('types')
        context['sort'] = self.request.GET.get('sort', 'date_debut')
        return context


class RapportEvenementsView(StaffRequiredMixin, TemplateView):
    """
    Rapports et statistiques des événements
    """
    template_name = 'evenements/rapports/evenements.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Période d'analyse
        date_debut = self.request.GET.get('date_debut')
        date_fin = self.request.GET.get('date_fin')
        
        if not date_debut:
            date_debut = (timezone.now() - timedelta(days=365)).date()
        else:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        
        if not date_fin:
            date_fin = timezone.now().date()
        else:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        
        # Statistiques générales
        evenements = Evenement.objects.filter(
            date_debut__date__range=[date_debut, date_fin]
        )
        
        stats_generales = evenements.aggregate(
            total_evenements=Count('id'),
            evenements_publies=Count('id', filter=Q(statut='publie')),
            evenements_annules=Count('id', filter=Q(statut='annule')),
            total_inscriptions=Sum('inscriptions__id'),
            participants_total=Sum('inscriptions__nombre_accompagnants', filter=Q(inscriptions__statut='confirmee')),
            revenus_total=Sum('inscriptions__montant_paye')
        )
        
        # Statistiques par type
        stats_par_type = TypeEvenement.objects.annotate(
            nb_evenements=Count('evenement', filter=Q(
                evenement__date_debut__date__range=[date_debut, date_fin]
            )),
            nb_participants=Sum('evenement__inscriptions__nombre_accompagnants', filter=Q(
                evenement__date_debut__date__range=[date_debut, date_fin],
                evenement__inscriptions__statut='confirmee'
            ))
        ).filter(nb_evenements__gt=0)
        
        # Évolution mensuelle
        evolution_mensuelle = []
        current_date = date_debut.replace(day=1)
        while current_date <= date_fin:
            next_month = current_date.replace(day=28) + timedelta(days=4)
            next_month = next_month.replace(day=1)
            
            stats_mois = evenements.filter(
                date_debut__date__range=[current_date, next_month - timedelta(days=1)]
            ).aggregate(
                nb_evenements=Count('id'),
                nb_participants=Sum('inscriptions__nombre_accompagnants', filter=Q(inscriptions__statut='confirmee')),
                revenus=Sum('inscriptions__montant_paye')
            )
            
            evolution_mensuelle.append({
                'mois': current_date.strftime('%Y-%m'),
                'mois_nom': current_date.strftime('%B %Y'),
                'nb_evenements': stats_mois['nb_evenements'] or 0,
                'nb_participants': stats_mois['nb_participants'] or 0,
                'revenus': stats_mois['revenus'] or 0,
            })
            
            current_date = next_month
        
        context.update({
            'date_debut': date_debut,
            'date_fin': date_fin,
            'stats_generales': stats_generales,
            'stats_par_type': stats_par_type,
            'evolution_mensuelle': evolution_mensuelle,
        })
        
        return context


class SessionListView(StaffRequiredMixin, ListView):
    """
    Liste des sessions d'un événement
    """
    model = SessionEvenement
    template_name = 'evenements/sessions/liste.html'
    context_object_name = 'sessions'
    
    def dispatch(self, request, *args, **kwargs):
        self.evenement = get_object_or_404(Evenement, pk=kwargs['evenement_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return SessionEvenement.objects.par_evenement(self.evenement).chronologique()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evenement'] = self.evenement
        return context


class SessionCreateView(StaffRequiredMixin, CreateView):
    """
    Création d'une session pour un événement
    """
    model = SessionEvenement
    form_class = SessionEvenementForm
    template_name = 'evenements/sessions/form.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.evenement = get_object_or_404(Evenement, pk=kwargs['evenement_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['evenement'] = self.evenement
        return kwargs
    
    def form_valid(self, form):
        form.instance.evenement_parent = self.evenement
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f"La session '{self.object.titre_session}' a été créée."
        )
        return response
    
    def get_success_url(self):
        return reverse('evenements:sessions_liste', kwargs={'evenement_pk': self.evenement.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evenement'] = self.evenement
        context['action'] = 'Créer'
        return context


class ExportBadgesView(StaffRequiredMixin, View):
    """
    Génération de badges pour les participants
    """
    
    def get(self, request, evenement_pk):
        evenement = get_object_or_404(Evenement, pk=evenement_pk)
        
        # Récupérer les participants confirmés
        inscriptions = InscriptionEvenement.objects.filter(
            evenement=evenement,
            statut__in=['confirmee', 'presente']
        ).select_related('membre').prefetch_related('accompagnants')
        
        return self._generer_badges_pdf(evenement, inscriptions)
    
    def _generer_badges_pdf(self, evenement, inscriptions):
        """Génère les badges en PDF"""
        from django.http import HttpResponse
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from io import BytesIO
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Dimensions des badges (4 par page)
        badge_width = 8*cm
        badge_height = 5*cm
        margin = 1*cm
        
        badges_per_row = 2
        badges_per_col = 4
        badges_per_page = badges_per_row * badges_per_col
        
        badge_count = 0
        
        for inscription in inscriptions:
            participants = [inscription.membre]
            participants.extend([acc for acc in inscription.accompagnants.all()])
            
            for participant in participants:
                if badge_count > 0 and badge_count % badges_per_page == 0:
                    p.showPage()
                
                # Position du badge sur la page
                col = badge_count % badges_per_row
                row = (badge_count % badges_per_page) // badges_per_row
                
                x = margin + col * (badge_width + margin)
                y = height - margin - (row + 1) * (badge_height + margin)
                
                # Dessiner le badge
                self._dessiner_badge(p, x, y, badge_width, badge_height, participant, evenement)
                
                badge_count += 1
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="badges_{evenement.reference}.pdf"'
        
        return response
    
    def _dessiner_badge(self, canvas, x, y, width, height, participant, evenement):
        """Dessine un badge individuel"""
        # Bordure
        canvas.rect(x, y, width, height)
        
        # Titre de l'événement
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredText(x + width/2, y + height - 1*cm, evenement.titre[:30])
        
        # Nom du participant
        canvas.setFont("Helvetica-Bold", 16)
        if hasattr(participant, 'prenom'):  # Membre
            nom_complet = f"{participant.prenom} {participant.nom}"
        else:  # Accompagnant
            nom_complet = f"{participant.prenom} {participant.nom}"
        
        canvas.drawCentredText(x + width/2, y + height/2, nom_complet[:25])
        
        # Date et lieu
        canvas.setFont("Helvetica", 10)
        date_str = evenement.date_debut.strftime('%d/%m/%Y')
        canvas.drawCentredText(x + width/2, y + 1*cm, f"{date_str} • {evenement.lieu[:20]}")


class GenererRecuView(LoginRequiredMixin, View):
    """
    Génération de reçu pour une inscription
    """
    
    def get(self, request, pk):
        inscription = get_object_or_404(InscriptionEvenement, pk=pk)
        
        # Vérifier les permissions
        if not request.user.is_staff and inscription.membre.utilisateur != request.user:
            raise Http404
        
        return self._generer_recu_pdf(inscription)
    
    def _generer_recu_pdf(self, inscription):
        """Génère un reçu en PDF"""
        from django.http import HttpResponse
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from io import BytesIO
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # En-tête
        p.setFont("Helvetica-Bold", 20)
        p.drawCentredText(width/2, height - 3*cm, "REÇU D'INSCRIPTION")
        
        # Informations
        y = height - 5*cm
        p.setFont("Helvetica", 12)
        
        lines = [
            f"Événement: {inscription.evenement.titre}",
            f"Date: {inscription.evenement.date_debut.strftime('%d/%m/%Y à %H:%M')}",
            f"Lieu: {inscription.evenement.lieu}",
            "",
            f"Participant: {inscription.membre.prenom} {inscription.membre.nom}",
            f"Email: {inscription.membre.email}",
            f"Date d'inscription: {inscription.date_inscription.strftime('%d/%m/%Y à %H:%M')}",
            "",
            f"Nombre d'accompagnants: {inscription.nombre_accompagnants}",
            f"Montant payé: {inscription.montant_paye}€",
            f"Mode de paiement: {inscription.mode_paiement.libelle if inscription.mode_paiement else 'Non spécifié'}",
        ]
        
        if inscription.reference_paiement:
            lines.append(f"Référence: {inscription.reference_paiement}")
        
        for line in lines:
            p.drawString(3*cm, y, line)
            y -= 0.6*cm
        
        # Pied de page
        p.setFont("Helvetica", 8)
        p.drawCentredText(width/2, 2*cm, f"Reçu généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="recu_{inscription.evenement.reference}_{inscription.membre.nom}.pdf"'
        
        return response


class ConfigurerRecurrenceView(StaffRequiredMixin, FormView):
    """
    Configuration de la récurrence d'un événement
    """
    form_class = EvenementRecurrenceForm
    template_name = 'evenements/recurrence/form.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.evenement = get_object_or_404(Evenement, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_initial(self):
        initial = super().get_initial()
        try:
            recurrence = self.evenement.recurrence
            initial.update({
                'frequence': recurrence.frequence,
                'intervalle_recurrence': recurrence.intervalle_recurrence,
                'jours_semaine_selection': recurrence.jours_semaine,
                'date_fin_recurrence': recurrence.date_fin_recurrence,
                'nombre_occurrences_max': recurrence.nombre_occurrences_max,
            })
        except EvenementRecurrence.DoesNotExist:
            pass
        return initial
    
    def form_valid(self, form):
        # Marquer l'événement comme récurrent
        self.evenement.est_recurrent = True
        self.evenement.save()
        
        # Créer ou mettre à jour la configuration
        recurrence, created = EvenementRecurrence.objects.get_or_create(
            evenement_parent=self.evenement,
            defaults={
                'frequence': form.cleaned_data['frequence'],
                'intervalle_recurrence': form.cleaned_data['intervalle_recurrence'],
                'jours_semaine': form.cleaned_data.get('jours_semaine', []),
                'date_fin_recurrence': form.cleaned_data.get('date_fin_recurrence'),
                'nombre_occurrences_max': form.cleaned_data.get('nombre_occurrences_max'),
            }
        )
        
        if not created:
            recurrence.frequence = form.cleaned_data['frequence']
            recurrence.intervalle_recurrence = form.cleaned_data['intervalle_recurrence']
            recurrence.jours_semaine = form.cleaned_data.get('jours_semaine', [])
            recurrence.date_fin_recurrence = form.cleaned_data.get('date_fin_recurrence')
            recurrence.nombre_occurrences_max = form.cleaned_data.get('nombre_occurrences_max')
            recurrence.save()
        
        messages.success(
            self.request,
            f"La récurrence pour '{self.evenement.titre}' a été configurée."
        )
        
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('evenements:detail', kwargs={'pk': self.evenement.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evenement'] = self.evenement
        return context


# =============================================================================
# VUES AJAX ET API
# =============================================================================

class CalculerTarifView(LoginRequiredMixin, AjaxRequiredMixin, View):
    """
    Calcul AJAX du tarif pour un membre
    """
    
    def get(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        
        try:
            membre = Membre.objects.get(utilisateur=request.user)
            tarif = evenement.calculer_tarif_membre(membre)
            
            return JsonResponse({
                'success': True,
                'tarif': float(tarif),
                'tarif_formate': f"{tarif}€",
                'est_payant': evenement.est_payant,
            })
        except Membre.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Membre non trouvé'
            })


class CheckPeutInscrireView(LoginRequiredMixin, AjaxRequiredMixin, View):
    """
    Vérification AJAX si un membre peut s'inscrire
    """
    
    def get(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        
        try:
            membre = Membre.objects.get(utilisateur=request.user)
            peut_inscrire, message = evenement.peut_s_inscrire(membre)
            
            return JsonResponse({
                'peut_inscrire': peut_inscrire,
                'message': message,
                'places_disponibles': evenement.places_disponibles,
                'est_complet': evenement.est_complet,
            })
        except Membre.DoesNotExist:
            return JsonResponse({
                'peut_inscrire': False,
                'message': 'Vous devez être membre pour vous inscrire',
                'places_disponibles': 0,
                'est_complet': True,
            })


class AutocompleteOrganisateursView(StaffRequiredMixin, AjaxRequiredMixin, View):
    """
    Autocomplétion AJAX pour les organisateurs
    """
    
    def get(self, request):
        query = request.GET.get('q', '')
        if len(query) < 2:
            return JsonResponse({'results': []})
        
        # Rechercher parmi les membres actifs
        organisateurs = Membre.objects.filter(
            Q(nom__icontains=query) | Q(prenom__icontains=query) | Q(email__icontains=query),
            deleted_at__isnull=True
        ).select_related('utilisateur')[:10]
        
        results = []
        for membre in organisateurs:
            results.append({
                'id': membre.utilisateur.id,
                'text': f"{membre.prenom} {membre.nom} ({membre.email})",
                'nom_complet': f"{membre.prenom} {membre.nom}",
                'email': membre.email,
            })
        
        return JsonResponse({'results': results})


class AutocompleteLieuxView(LoginRequiredMixin, AjaxRequiredMixin, View):
    """
    Autocomplétion AJAX pour les lieux
    """
    
    def get(self, request):
        query = request.GET.get('q', '')
        if len(query) < 2:
            return JsonResponse({'results': []})
        
        # Rechercher dans les lieux existants
        lieux = Evenement.objects.filter(
            lieu__icontains=query
        ).values_list('lieu', flat=True).distinct()[:10]
        
        results = [{'id': lieu, 'text': lieu} for lieu in lieux]
        
        return JsonResponse({'results': results})


class PromouvoirListeAttenteView(StaffRequiredMixin, AjaxRequiredMixin, View):
    """
    Promotion AJAX depuis la liste d'attente
    """
    
    def post(self, request, pk):
        inscription = get_object_or_404(InscriptionEvenement, pk=pk)
        
        if inscription.statut == 'liste_attente':
            # Vérifier s'il y a des places disponibles
            if inscription.evenement.places_disponibles > 0:
                inscription.statut = 'en_attente'
                inscription.date_limite_confirmation = timezone.now() + timezone.timedelta(
                    hours=inscription.evenement.delai_confirmation
                )
                inscription.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Inscription promue avec succès',
                    'nouveau_statut': inscription.get_statut_display()
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Aucune place disponible'
                })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Cette inscription n\'est pas en liste d\'attente'
            })


# =============================================================================
# VUES PUBLIQUES (SANS AUTHENTIFICATION)
# =============================================================================

class EvenementsPublicsView(ListView):
    """
    Vue publique des événements (sans authentification requise)
    """
    model = Evenement
    template_name = 'evenements/publics/liste.html'
    context_object_name = 'evenements'
    paginate_by = 12
    
    def get_queryset(self):
        # CORRECTION : Utiliser les nouvelles méthodes du gestionnaire
        return Evenement.objects.publies().a_venir().select_related('type_evenement', 'organisateur')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['types_evenements'] = TypeEvenement.objects.par_ordre_affichage()
        return context


class EvenementPublicDetailView(DetailView):
    """
    Détail public d'un événement
    """
    model = Evenement
    template_name = 'evenements/publics/detail.html'
    context_object_name = 'evenement'
    
    def get_queryset(self):
        # CORRECTION : Utiliser les nouvelles méthodes du gestionnaire
        return Evenement.objects.publies()


class CalendrierPublicView(TemplateView):
    """
    Calendrier public des événements
    """
    template_name = 'evenements/publics/calendrier.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # CORRECTION : Utiliser les nouvelles méthodes du gestionnaire
        evenements = Evenement.objects.publies().a_venir()
        
        # Sérialiser pour le calendrier JavaScript
        evenements_json = []
        for evt in evenements:
            evenements_json.append({
                'id': evt.id,
                'title': evt.titre,
                'start': evt.date_debut.isoformat(),
                'end': evt.date_fin.isoformat(),
                'url': reverse('evenements:evenement_public_detail', kwargs={'pk': evt.pk})
            })
        
        context['evenements_json'] = json.dumps(evenements_json)
        return context

# =============================================================================
# VUES D'EXPORT ET IMPORT
# =============================================================================

class ExportCalendrierView(LoginRequiredMixin, View):
    """
    Export calendrier iCal
    """
    
    def get(self, request):
        from django.http import HttpResponse
        import uuid
        from datetime import datetime
        
        # Récupérer les événements de l'utilisateur
        try:
            membre = Membre.objects.get(utilisateur=request.user)
            mes_evenements = InscriptionEvenement.objects.filter(
                membre=membre,
                statut__in=['confirmee', 'presente']
            ).select_related('evenement')
        except Membre.DoesNotExist:
            mes_evenements = []
        
        # Générer le contenu iCal
        ical_content = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//Association//Gestion Evenements//FR",
            "CALSCALE:GREGORIAN",
            "METHOD:PUBLISH",
            "X-WR-CALNAME:Mes Événements",
            "X-WR-TIMEZONE:Europe/Paris",
        ]
        
        for inscription in mes_evenements:
            evenement = inscription.evenement
            
            # Format de date iCal (UTC)
            dt_start = evenement.date_debut.strftime('%Y%m%dT%H%M%SZ')
            dt_end = evenement.date_fin.strftime('%Y%m%dT%H%M%SZ') if evenement.date_fin else dt_start
            
            ical_content.extend([
                "BEGIN:VEVENT",
                f"UID:{evenement.reference}@association.local",
                f"DTSTART:{dt_start}",
                f"DTEND:{dt_end}",
                f"SUMMARY:{evenement.titre}",
                f"DESCRIPTION:{evenement.description[:200]}...",
                f"LOCATION:{evenement.lieu}",
                f"ORGANIZER:CN={evenement.organisateur.get_full_name()}",
                f"STATUS:CONFIRMED",
                "END:VEVENT",
            ])
        
        ical_content.append("END:VCALENDAR")
        
        response = HttpResponse('\r\n'.join(ical_content), content_type='text/calendar; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="mes_evenements.ics"'
        
        return response


class ImportEvenementsView(StaffRequiredMixin, FormView):
    """
    Import d'événements depuis un fichier
    """
    form_class = ExportEvenementsForm  # Réutiliser pour le choix de format
    template_name = 'evenements/import/form.html'
    success_url = reverse_lazy('evenements:liste')
    
    def form_valid(self, form):
        fichier = self.request.FILES.get('fichier')
        if not fichier:
            messages.error(self.request, "Aucun fichier sélectionné.")
            return self.form_invalid(form)
        
        try:
            if fichier.name.endswith('.csv'):
                resultat = self._importer_csv(fichier)
            elif fichier.name.endswith('.xlsx'):
                resultat = self._importer_excel(fichier)
            else:
                messages.error(self.request, "Format de fichier non supporté.")
                return self.form_invalid(form)
            
            messages.success(
                self.request,
                f"Import terminé : {resultat['succes']} événements importés, {resultat['erreurs']} erreurs."
            )
            
        except Exception as e:
            messages.error(self.request, f"Erreur lors de l'import : {str(e)}")
            return self.form_invalid(form)
        
        return super().form_valid(form)
    
    def _importer_csv(self, fichier):
        """Import depuis un fichier CSV"""
        import csv
        import io
        
        content = fichier.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        
        succes = 0
        erreurs = 0
        
        for row in reader:
            try:
                # Créer l'événement depuis les données CSV
                # (implémentation simplifiée)
                evenement = Evenement(
                    titre=row['titre'],
                    description=row.get('description', ''),
                    lieu=row['lieu'],
                    # ... autres champs
                )
                evenement.full_clean()
                evenement.save()
                succes += 1
            except Exception:
                erreurs += 1
        
        return {'succes': succes, 'erreurs': erreurs}
    
    def _importer_excel(self, fichier):
        """Import depuis un fichier Excel"""
        # Implémentation similaire avec openpyxl
        return {'succes': 0, 'erreurs': 0}


# =============================================================================
# VUES DE MAINTENANCE ET UTILITAIRES
# =============================================================================

class NettoyerInscriptionsView(StaffRequiredMixin, View):
    """
    Nettoyage des inscriptions expirées
    """
    
    def post(self, request):
        count = InscriptionEvenement.objects.nettoyer_inscriptions_expirees()
        messages.success(
            request,
            f"{count} inscription(s) expirée(s) ont été nettoyées."
        )
        return redirect('evenements:dashboard')


class ValidationMasseView(StaffRequiredMixin, FormView):
    """
    Validation en masse d'événements
    """
    template_name = 'evenements/validation/masse.html'
    
    def post(self, request):
        action = request.POST.get('action')
        evenement_ids = request.POST.getlist('evenements')
        commentaire = request.POST.get('commentaire', '')
        
        if not evenement_ids:
            messages.error(request, "Aucun événement sélectionné.")
            return redirect('evenements:validation_liste')
        
        count = 0
        validations = ValidationEvenement.objects.filter(
            evenement__id__in=evenement_ids,
            statut_validation='en_attente'
        )
        
        for validation in validations:
            if action == 'approuver':
                validation.approuver(request.user, commentaire)
                count += 1
            elif action == 'refuser' and commentaire:
                validation.refuser(request.user, commentaire)
                count += 1
        
        if count > 0:
            messages.success(
                request,
                f"{count} événement(s) {action}(s) avec succès."
            )
        else:
            messages.warning(request, "Aucun événement traité.")
        
        return redirect('evenements:validation_liste')


# =============================================================================
# VUES DE CORBEILLE (SUPPRESSION LOGIQUE)
# =============================================================================

class CorbeilleEvenementsView(StaffRequiredMixin, ListView):
    """
    Liste des événements supprimés (corbeille)
    """
    model = Evenement
    template_name = 'evenements/corbeille/evenements.html'
    context_object_name = 'evenements'
    paginate_by = 20
    
    def get_queryset(self):
        return Evenement.objects.only_deleted().select_related('type_evenement', 'organisateur')


class RestaurerEvenementView(StaffRequiredMixin, View):
    """
    Restauration d'un événement supprimé
    """
    
    def post(self, request, pk):
        try:
            evenement = Evenement.objects.only_deleted().get(pk=pk)
            evenement.deleted_at = None
            evenement.save()
            
            messages.success(
                request,
                f"L'événement '{evenement.titre}' a été restauré."
            )
        except Evenement.DoesNotExist:
            messages.error(request, "Événement non trouvé.")
        
        return redirect('evenements:corbeille_evenements')

class ConfirmationPubliqueView(View):
    """
    Vue de confirmation publique - placeholder
    """
    def get(self, request, code):
        # TODO: Implémenter la logique de confirmation publique
        return HttpResponse("Confirmation publique - À implémenter")


# Gestion d'erreurs simplifiées
def evenement_404(request, exception=None):
    """Vue d'erreur 404 personnalisée pour les événements"""
    from django.shortcuts import render
    return render(request, 'core/errors/404.html', status=404)


def evenement_500(request):
    """Vue d'erreur 500 personnalisée pour les événements"""
    from django.shortcuts import render
    return render(request, 'core/errors/500.html', status=500)


# =============================================================================
# VUES GESTION ÉVÉNEMENTS AVANCÉES
# =============================================================================

class EvenementDuplicateView(StaffRequiredMixin, View):
    """
    Duplication d'un événement
    """
    
    def post(self, request, pk):
        evenement_original = get_object_or_404(Evenement, pk=pk)
        
        # Créer une copie
        evenement_copie = Evenement.objects.get(pk=pk)
        evenement_copie.pk = None
        evenement_copie.titre = f"Copie de {evenement_original.titre}"
        evenement_copie.statut = 'brouillon'
        evenement_copie.reference = None  # Sera régénéré
        evenement_copie.save()
        
        messages.success(
            request,
            f"L'événement '{evenement_copie.titre}' a été dupliqué."
        )
        
        return redirect('evenements:detail', pk=evenement_copie.pk)


class PublierEvenementView(StaffRequiredMixin, View):
    """
    Publication d'un événement
    """
    
    def post(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        
        if evenement.statut == 'brouillon':
            evenement.statut = 'publie'
            evenement.save()
            
            messages.success(
                request,
                f"L'événement '{evenement.titre}' a été publié."
            )
        else:
            messages.warning(
                request,
                "Cet événement ne peut pas être publié dans son état actuel."
            )
        
        return redirect('evenements:detail', pk=evenement.pk)


class AnnulerEvenementView(StaffRequiredMixin, FormView):
    """
    Annulation d'un événement
    """
    template_name = 'evenements/annuler_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evenement'] = get_object_or_404(Evenement, pk=self.kwargs['pk'])
        return context
    
    def post(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        raison = request.POST.get('raison', '')
        
        evenement.statut = 'annule'
        evenement.save()
        
        # TODO: Notifier les inscrits de l'annulation
        
        messages.success(
            request,
            f"L'événement '{evenement.titre}' a été annulé."
        )
        
        return redirect('evenements:liste')


class ReporterEvenementView(StaffRequiredMixin, FormView):
    """
    Report d'un événement
    """
    template_name = 'evenements/reporter_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evenement'] = get_object_or_404(Evenement, pk=self.kwargs['pk'])
        return context
    
    def post(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        nouvelle_date = request.POST.get('nouvelle_date')
        
        if nouvelle_date:
            evenement.statut = 'reporte'
            # TODO: Logique de report avec nouvelle date
            evenement.save()
            
            messages.success(
                request,
                f"L'événement '{evenement.titre}' a été reporté."
            )
        
        return redirect('evenements:detail', pk=evenement.pk)


class GenererOccurrencesView(StaffRequiredMixin, View):
    """
    Génération d'occurrences pour événements récurrents
    """
    
    def post(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        
        if not evenement.est_recurrent:
            messages.error(request, "Cet événement n'est pas récurrent.")
            return redirect('evenements:detail', pk=pk)
        
        # TODO: Logique de génération des occurrences
        count = 0  # Nombre d'occurrences générées
        
        messages.success(
            request,
            f"{count} occurrence(s) générée(s) pour '{evenement.titre}'."
        )
        
        return redirect('evenements:detail', pk=pk)


# =============================================================================
# VUES SESSIONS
# =============================================================================

class SessionDetailView(LoginRequiredMixin, DetailView):
    """
    Détail d'une session
    """
    model = SessionEvenement
    template_name = 'evenements/sessions/detail.html'
    context_object_name = 'session'


class SessionUpdateView(StaffRequiredMixin, UpdateView):
    """
    Modification d'une session
    """
    model = SessionEvenement
    form_class = SessionEvenementForm
    template_name = 'evenements/sessions/form.html'
    
    def get_success_url(self):
        return reverse('evenements:session_detail', kwargs={'pk': self.object.pk})


class SessionDeleteView(StaffRequiredMixin, DeleteView):
    """
    Suppression d'une session
    """
    model = SessionEvenement
    template_name = 'evenements/sessions/confirm_delete.html'
    
    def get_success_url(self):
        return reverse('evenements:sessions_liste', 
                      kwargs={'evenement_pk': self.object.evenement_parent.pk})


# =============================================================================
# VUES INSCRIPTIONS AVANCÉES
# =============================================================================

class InscriptionUpdateView(LoginRequiredMixin, UpdateView):
    """
    Modification d'une inscription
    """
    model = InscriptionEvenement
    form_class = InscriptionEvenementForm
    template_name = 'evenements/inscription/form.html'
    
    def get_queryset(self):
        queryset = InscriptionEvenement.objects.all()
        if not self.request.user.is_staff:
            try:
                membre = Membre.objects.get(utilisateur=self.request.user)
                queryset = queryset.filter(membre=membre)
            except Membre.DoesNotExist:
                queryset = queryset.none()
        return queryset
    
    def get_success_url(self):
        return reverse('evenements:inscription_detail', kwargs={'pk': self.object.pk})


class AccompagnantListView(LoginRequiredMixin, ListView):
    """
    Liste des accompagnants d'une inscription
    """
    model = AccompagnantInvite
    template_name = 'evenements/accompagnants/liste.html'
    context_object_name = 'accompagnants'
    
    def dispatch(self, request, *args, **kwargs):
        self.inscription = get_object_or_404(InscriptionEvenement, pk=kwargs['inscription_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        return AccompagnantInvite.objects.filter(inscription=self.inscription)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inscription'] = self.inscription
        return context


class AccompagnantCreateView(LoginRequiredMixin, CreateView):
    """
    Ajout d'un accompagnant
    """
    model = AccompagnantInvite
    form_class = AccompagnantForm
    template_name = 'evenements/accompagnants/form.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.inscription = get_object_or_404(InscriptionEvenement, pk=kwargs['inscription_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        form.instance.inscription = self.inscription
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('evenements:accompagnants_liste', 
                      kwargs={'inscription_pk': self.inscription.pk})


class AccompagnantDetailView(LoginRequiredMixin, DetailView):
    """
    Détail d'un accompagnant
    """
    model = AccompagnantInvite
    template_name = 'evenements/accompagnants/detail.html'
    context_object_name = 'accompagnant'


class AccompagnantUpdateView(LoginRequiredMixin, UpdateView):
    """
    Modification d'un accompagnant
    """
    model = AccompagnantInvite
    form_class = AccompagnantForm
    template_name = 'evenements/accompagnants/form.html'
    
    def get_success_url(self):
        return reverse('evenements:accompagnant_detail', kwargs={'pk': self.object.pk})


class AccompagnantDeleteView(LoginRequiredMixin, DeleteView):
    """
    Suppression d'un accompagnant
    """
    model = AccompagnantInvite
    template_name = 'evenements/accompagnants/confirm_delete.html'
    
    def get_success_url(self):
        return reverse('evenements:accompagnants_liste', 
                      kwargs={'inscription_pk': self.object.inscription.pk})


class PaiementInscriptionView(LoginRequiredMixin, FormView):
    """
    Paiement d'une inscription
    """
    form_class = PaiementInscriptionForm
    template_name = 'evenements/inscription/paiement.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.inscription = get_object_or_404(InscriptionEvenement, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['inscription'] = self.inscription
        return kwargs
    
    def form_valid(self, form):
        # TODO: Logique d'enregistrement du paiement
        messages.success(self.request, "Paiement enregistré avec succès.")
        return redirect('evenements:inscription_detail', pk=self.inscription.pk)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inscription'] = self.inscription
        return context


# =============================================================================
# VUES VALIDATION AVANCÉES
# =============================================================================

class DemanderModificationsView(StaffRequiredMixin, FormView):
    """
    Demander des modifications sur un événement
    """
    form_class = ValidationEvenementForm
    template_name = 'evenements/validation/demander_modifications.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.validation = get_object_or_404(ValidationEvenement, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        modifications = form.cleaned_data['commentaire_validation']
        self.validation.demander_modifications(self.request.user, modifications)
        
        messages.success(
            self.request,
            "Demande de modifications envoyée à l'organisateur."
        )
        
        return redirect('evenements:validation_liste')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['validation'] = self.validation
        return context


class ValidationStatsView(StaffRequiredMixin, TemplateView):
    """
    Statistiques de validation
    """
    template_name = 'evenements/validation/statistiques.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistiques globales
        context['stats'] = ValidationEvenement.objects.aggregate(
            total=Count('id'),
            en_attente=Count('id', filter=Q(statut_validation='en_attente')),
            approuvees=Count('id', filter=Q(statut_validation='approuve')),
            refusees=Count('id', filter=Q(statut_validation='refuse'))
        )
        
        return context


# =============================================================================
# VUES TYPES D'ÉVÉNEMENTS AVANCÉES
# =============================================================================

class TypeEvenementDetailView(LoginRequiredMixin, DetailView):
    """
    Détail d'un type d'événement
    """
    model = TypeEvenement
    # CORRECTION : Utiliser un template générique ou existant
    template_name = 'evenements/types/liste.html'  # Réutiliser le template de liste temporairement
    context_object_name = 'type_evenement'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['evenements'] = Evenement.objects.filter(
            type_evenement=self.object
        ).order_by('-date_debut')[:10]
        return context


class TypeEvenementUpdateView(StaffRequiredMixin, UpdateView):
    """
    Modification d'un type d'événement
    """
    model = TypeEvenement
    template_name = 'evenements/form.html'  # au lieu de 'evenements/types/form.html'
    fields = ['libelle', 'description', 'couleur_affichage', 'permet_accompagnants', 'necessite_validation']
    success_url = reverse_lazy('evenements:types_liste')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f"Le type d'événement '{self.object.libelle}' a été modifié."
        )
        return response


class TypeEvenementDeleteView(StaffRequiredMixin, DeleteView):
    """
    Suppression d'un type d'événement
    """
    model = TypeEvenement
    template_name = 'evenements/types/confirm_delete.html'
    success_url = reverse_lazy('evenements:types_liste')


# =============================================================================
# VUES EXPORTS AVANCÉES
# =============================================================================

class ExportEvenementsView(StaffRequiredMixin, FormView):
    """
    Export général des événements
    """
    form_class = ExportEvenementsForm
    template_name = 'evenements/export/evenements.html'
    
    def form_valid(self, form):
        # TODO: Logique d'export général
        return HttpResponse("Export en cours de développement")


class ExportEvenementView(StaffRequiredMixin, View):
    """
    Export d'un événement spécifique
    """
    
    def get(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        format_export = request.GET.get('format', 'pdf')
        
        # TODO: Logique d'export selon format
        return HttpResponse(f"Export {format_export} de {evenement.titre}")


class ExportInscriptionsView(StaffRequiredMixin, View):
    """
    Export général des inscriptions
    """
    
    def get(self, request):
        # TODO: Logique d'export des inscriptions
        return HttpResponse("Export inscriptions en cours de développement")


class ExportInscriptionsEvenementView(StaffRequiredMixin, View):
    """
    Export des inscriptions d'un événement (déjà implémenté comme ExportInscritsView)
    """
    
    def get(self, request, evenement_pk):
        return ExportInscritsView.as_view()(request, evenement_pk=evenement_pk)


class ExportEvenementCalendrierView(LoginRequiredMixin, View):
    """
    Export iCal d'un événement spécifique
    """
    
    def get(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        
        # Générer le contenu iCal pour un événement
        ical_content = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//Association//Gestion Evenements//FR",
            "BEGIN:VEVENT",
            f"UID:{evenement.reference}@association.local",
            f"DTSTART:{evenement.date_debut.strftime('%Y%m%dT%H%M%SZ')}",
            f"DTEND:{evenement.date_fin.strftime('%Y%m%dT%H%M%SZ') if evenement.date_fin else evenement.date_debut.strftime('%Y%m%dT%H%M%SZ')}",
            f"SUMMARY:{evenement.titre}",
            f"DESCRIPTION:{evenement.description[:200]}...",
            f"LOCATION:{evenement.lieu}",
            "END:VEVENT",
            "END:VCALENDAR"
        ]
        
        response = HttpResponse('\r\n'.join(ical_content), content_type='text/calendar')
        response['Content-Disposition'] = f'attachment; filename="{evenement.reference}.ics"'
        
        return response


# =============================================================================
# VUES RAPPORTS AVANCÉES
# =============================================================================

class RapportDashboardView(LoginRequiredMixin, TemplateView):  # CORRECTION : Enlever StaffRequiredMixin
    """
    Dashboard des rapports - Accessible aux membres connectés
    """
    template_name = 'evenements/rapports/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Données de base pour tous les utilisateurs connectés
        context['evenements_recents'] = Evenement.objects.publies()[:5]
        
        # Données détaillées uniquement pour les staff
        if self.request.user.is_staff:
            context.update({
                'stats_generales': {
                    'total_evenements': Evenement.objects.count(),
                    'evenements_publies': Evenement.objects.publies().count(),
                    'inscriptions_total': InscriptionEvenement.objects.count(),
                }
            })
        
        return context


class RapportFrequentationView(StaffRequiredMixin, TemplateView):
    """
    Rapport de fréquentation
    """
    template_name = 'evenements/rapports/frequentation.html'


class RapportRevenusView(StaffRequiredMixin, TemplateView):
    """
    Rapport des revenus
    """
    template_name = 'evenements/rapports/revenus.html'


class RapportParticipationMembresView(StaffRequiredMixin, TemplateView):
    """
    Rapport de participation des membres
    """
    template_name = 'evenements/rapports/participation_membres.html'


class RapportFideliteMembresView(StaffRequiredMixin, TemplateView):
    """
    Rapport de fidélité des membres
    """
    template_name = 'evenements/rapports/fidelite_membres.html'


class RapportOrganisateursView(StaffRequiredMixin, TemplateView):
    """
    Rapport des organisateurs
    """
    template_name = 'evenements/rapports/organisateurs.html'


# =============================================================================
# VUES AJAX AVANCÉES
# =============================================================================

class CalculerMontantInscriptionView(LoginRequiredMixin, AjaxRequiredMixin, View):
    """
    Calcul AJAX du montant d'inscription
    """
    
    def get(self, request, pk):
        inscription = get_object_or_404(InscriptionEvenement, pk=pk)
        
        return JsonResponse({
            'montant_total': float(inscription.calculer_montant_total()),
            'montant_restant': float(inscription.montant_restant),
            'est_payee': inscription.est_payee
        })


class NotificationsInscriptionsView(AjaxRequiredMixin, View):
    """
    Notifications temps réel pour inscriptions
    """
    
    def get(self, request):
        # TODO: Logique de notifications temps réel
        return JsonResponse({
            'notifications': [],
            'count': 0
        })


class NotificationsValidationsView(AjaxRequiredMixin, View):
    """
    Notifications temps réel pour validations
    """
    
    def get(self, request):
        if not request.user.is_staff:
            return JsonResponse({'error': 'Accès refusé'}, status=403)
        
        validations_urgentes = ValidationEvenement.objects.urgentes().count()
        
        return JsonResponse({
            'validations_urgentes': validations_urgentes,
            'message': f"{validations_urgentes} validation(s) urgente(s)" if validations_urgentes > 0 else ""
        })


class PreviewEvenementView(AjaxRequiredMixin, View):
    """
    Prévisualisation d'événement
    """
    
    def post(self, request):
        # TODO: Logique de prévisualisation
        return JsonResponse({
            'html': '<div>Prévisualisation en cours de développement</div>'
        })


# =============================================================================
# VUES CORBEILLE AVANCÉES
# =============================================================================

class CorbeilleInscriptionsView(StaffRequiredMixin, ListView):
    """
    Corbeille des inscriptions supprimées
    """
    model = InscriptionEvenement
    template_name = 'evenements/corbeille/inscriptions.html'
    context_object_name = 'inscriptions'
    
    def get_queryset(self):
        return InscriptionEvenement.objects.only_deleted()


class RestaurerInscriptionView(StaffRequiredMixin, View):
    """
    Restauration d'une inscription supprimée
    """
    
    def post(self, request, pk):
        try:
            inscription = InscriptionEvenement.objects.only_deleted().get(pk=pk)
            inscription.deleted_at = None
            inscription.save()
            
            messages.success(
                request,
                "L'inscription a été restaurée."
            )
        except InscriptionEvenement.DoesNotExist:
            messages.error(request, "Inscription non trouvée.")
        
        return redirect('evenements:corbeille_inscriptions')


class SupprimerDefinitivementEvenementView(StaffRequiredMixin, View):
    """
    Suppression définitive d'un événement
    """
    
    def post(self, request, pk):
        try:
            evenement = Evenement.objects.only_deleted().get(pk=pk)
            titre = evenement.titre
            evenement.delete(hard=True)  # Suppression physique
            
            messages.success(
                request,
                f"L'événement '{titre}' a été supprimé définitivement."
            )
        except Evenement.DoesNotExist:
            messages.error(request, "Événement non trouvé.")
        
        return redirect('evenements:corbeille_evenements')


# =============================================================================
# API ET FLUX RSS
# =============================================================================

class APIEvenementsPublicsView(View):
    """
    API JSON des événements publics
    """
    
    def get(self, request):
        evenements = Evenement.objects.publies().a_venir()[:10]
        
        data = []
        for evenement in evenements:
            data.append({
                'id': evenement.id,
                'titre': evenement.titre,
                'date_debut': evenement.date_debut.isoformat(),
                'lieu': evenement.lieu,
                'places_disponibles': evenement.places_disponibles
            })
        
        return JsonResponse({'evenements': data})


class APIEvenementDetailView(View):
    """
    API JSON détail d'un événement
    """
    
    def get(self, request, pk):
        evenement = get_object_or_404(Evenement.objects.publies(), pk=pk)
        
        data = {
            'id': evenement.id,
            'titre': evenement.titre,
            'description': evenement.description,
            'date_debut': evenement.date_debut.isoformat(),
            'date_fin': evenement.date_fin.isoformat() if evenement.date_fin else None,
            'lieu': evenement.lieu,
            'capacite_max': evenement.capacite_max,
            'places_disponibles': evenement.places_disponibles,
            'est_payant': evenement.est_payant,
            'tarif_membre': float(evenement.tarif_membre) if evenement.est_payant else 0
        }
        
        return JsonResponse(data)


class APIPlacesDisponiblesView(View):
    """
    API pour vérifier les places disponibles
    """
    
    def get(self, request, pk):
        evenement = get_object_or_404(Evenement, pk=pk)
        
        return JsonResponse({
            'places_disponibles': evenement.places_disponibles,
            'est_complet': evenement.est_complet,
            'taux_occupation': evenement.taux_occupation
        })


class EvenementsFeedView(Feed):
    """
    Flux RSS des événements
    """
    title = "Événements de l'Association"
    link = "/evenements/"
    description = "Derniers événements publiés"
    
    def items(self):
        return Evenement.objects.publies().a_venir()[:10]
    
    def item_title(self, item):
        return item.titre
    
    def item_description(self, item):
        return item.description[:200] + "..."
    
    def item_link(self, item):
        return reverse('evenements:evenement_public_detail', args=[item.pk])


class EvenementsAtomFeedView(EvenementsFeedView):
    """
    Flux Atom des événements
    """
    feed_type = Atom1Feed
    subtitle = EvenementsFeedView.description


# =============================================================================
# VUES UTILITAIRES ET AIDE
# =============================================================================

class AideEvenementsView(LoginRequiredMixin, TemplateView):
    """
    Page d'aide pour les événements
    """
    template_name = 'evenements/aide/index.html'


class DocumentationView(LoginRequiredMixin, TemplateView):
    """
    Documentation du module événements
    """
    template_name = 'evenements/aide/documentation.html'


class DownloadImportTemplateView(StaffRequiredMixin, View):
    """
    Téléchargement du template d'import
    """
    
    def get(self, request):
        # Créer un fichier Excel template
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Événements"
        
        # Headers
        headers = [
            'titre', 'description', 'type_evenement', 'date_debut',
            'date_fin', 'lieu', 'capacite_max', 'est_payant',
            'tarif_membre', 'organisateur_email'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Exemple de ligne
        exemple = [
            'Événement exemple', 'Description de l\'événement',
            'Formation', '2024-12-31 14:00', '2024-12-31 17:00',
            'Salle de conférence', '50', 'OUI', '25.00',
            'organisateur@example.com'
        ]
        
        for col, value in enumerate(exemple, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Sauvegarder en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_import_evenements.xlsx"'
        
        return response


# =============================================================================
# VUES CONFIGURATION ET PARAMÉTRAGE
# =============================================================================

class ConfigurationEvenementsView(StaffRequiredMixin, TemplateView):
    """
    Configuration du module événements
    """
    template_name = 'evenements/admin/configuration.html'


class ParametresEvenementsView(StaffRequiredMixin, TemplateView):
    """
    Paramètres du module événements
    """
    template_name = 'evenements/admin/parametres.html'


# =============================================================================
# VUES MAINTENANCE
# =============================================================================

class RecalculerStatsView(StaffRequiredMixin, View):
    """
    Recalcul des statistiques
    """
    
    def post(self, request):
        # TODO: Logique de recalcul des statistiques
        cache.clear()  # Vider le cache
        
        messages.success(
            request,
            "Les statistiques ont été recalculées."
        )
        
        return redirect('evenements:dashboard')


# =============================================================================
# WEBHOOKS ET INTÉGRATIONS
# =============================================================================

@method_decorator(csrf_exempt, name='dispatch')
class WebhookPaiementView(View):
    """
    Webhook pour les notifications de paiement
    """
    
    def post(self, request):
        # TODO: Traitement des webhooks de paiement
        return JsonResponse({'status': 'ok'})


class IntegrationCalendrierView(StaffRequiredMixin, TemplateView):
    """
    Intégration avec calendriers externes
    """
    template_name = 'evenements/integration/calendrier.html'


# =============================================================================
# TÂCHES CRON (POUR LES TESTS)
# =============================================================================

class CronRappelsView(View):
    """
    Tâche cron pour l'envoi des rappels
    """
    
    def get(self, request):
        if not request.META.get('HTTP_X_CRON_TOKEN') == 'secret_token':
            return HttpResponse('Unauthorized', status=401)
        
        # TODO: Logique d'envoi des rappels
        return HttpResponse('Rappels traités')


class CronNettoyageView(View):
    """
    Tâche cron pour le nettoyage
    """
    
    def get(self, request):
        if not request.META.get('HTTP_X_CRON_TOKEN') == 'secret_token':
            return HttpResponse('Unauthorized', status=401)
        
        count = InscriptionEvenement.objects.nettoyer_inscriptions_expirees()
        return HttpResponse(f'{count} inscriptions nettoyées')


class CronStatistiquesView(View):
    """
    Tâche cron pour la mise à jour des statistiques
    """
    
    def get(self, request):
        if not request.META.get('HTTP_X_CRON_TOKEN') == 'secret_token':
            return HttpResponse('Unauthorized', status=401)
        
        # TODO: Logique de mise à jour des statistiques
        return HttpResponse('Statistiques mises à jour')


# =============================================================================
# WIDGETS POUR INTÉGRATION
# =============================================================================

class WidgetProchainsEvenementsView(View):
    """
    Widget des prochains événements
    """
    
    def get(self, request):
        evenements = Evenement.objects.publies().a_venir()[:5]
        
        # Retourner en JSON pour intégration
        data = []
        for evenement in evenements:
            data.append({
                'titre': evenement.titre,
                'date': evenement.date_debut.strftime('%d/%m/%Y'),
                'lieu': evenement.lieu
            })
        
        return JsonResponse({'evenements': data})


class WidgetCalendrierMiniView(View):
    """
    Widget calendrier mini
    """
    
    def get(self, request):
        # TODO: Logique du calendrier mini
        return JsonResponse({
            'html': '<div>Calendrier mini en cours de développement</div>'
        })


# =============================================================================
# GESTION D'ERREURS SPÉCIALISÉE
# =============================================================================

def evenement_404(request, exception=None):
    """Vue d'erreur 404 personnalisée pour les événements"""
    return render(request, 'evenements/errors/404.html', {
        'message': 'Événement non trouvé'
    }, status=404)


def evenement_500(request):
    """Vue d'erreur 500 personnalisée pour les événements"""
    return render(request, 'evenements/errors/500.html', {
        'message': 'Erreur serveur dans le module événements'
    }, status=500)

# =============================================================================
# VUES D'ERREUR SPÉCIFIQUES
# =============================================================================

def evenement_404(request, exception):
    """Vue d'erreur 404 personnalisée pour les événements"""
    return render(request, 'evenements/errors/404.html', status=404)


def evenement_500(request):
    """Vue d'erreur 500 personnalisée pour les événements"""
    return render(request, 'evenements/errors/500.html', status=500)